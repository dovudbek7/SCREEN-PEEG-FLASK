"""7シート Excel 出力 (出張精算 承認チェックシート).

データ契約 §8 の 7 シートを openpyxl で書き出す. 入力は checksheet.py の
build_check_sheet() が返す dict (下記 INPUT CONTRACT). dict キーが欠けても
.get(..., []) で安全に空シートを出す.

INPUT CONTRACT (check_sheet dict) — checksheet.build_check_sheet と一致:
  {
    'reports':        [ {サマリ行 dict} ... ],   # 1 伝票 = 1 行 (総合判定含む)
    'legs':           [ {明細行 dict} ... ],      # 1 明細レッグ = 1 行
    'trip_labor':     [ {出張実態・労務 行 dict} ... ],
    'amount_receipt': [ {金額・領収書 行 dict} ... ],
    'duplicates':     [ {二重申請 行 dict} ... ],
    'approval':       [ {承認ルート 行 dict} ... ],
    'unmatched':      [ {未突合・要対応 行 dict} ... ],
    'banners':        [str, ...],                  # 既知欠落 等のバナー行
  }
各行 dict のキー = 列見出し. 列順は下の SHEETS で固定定義し, dict に欠ける
キーは空欄. dict に余分なキーがあれば末尾に追補列として出力する (取りこぼし防止).

ステータス語彙 (models): OK / 要確認 / NG / 未突合 / 未確認(規程未提供) /
未確認(勤怠データ欠落). 値に応じてセルを色分けする.
"""
from __future__ import annotations

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models import OK, NEEDS_CHECK, NG, UNMATCHED, MULTI, RULE_MISSING, ATT_MISSING

# ---------------------------------------------------------------------------
# シート定義 (spec §6 の 7 シート). (シート名 -> (dict キー, 固定列順))
# 列順は checksheet.build_check_sheet の各行 dict の挿入順に一致させてある.
# ---------------------------------------------------------------------------
SHEETS = [
    ("01_一次承認チェック", "primary", [
        "No.", "伝票No.", "入力者名", "社員番号", "所属", "承認状態",
        "出張期間", "出張先", "取引先", "出張実態",
        # 2026-07-31 客先依頼: 定時外の移動時間を日単位で表示するため 明細日付 を追加.
        # 1伝票が日数分の行に展開される. 識別列 (伝票No./入力者名等) は全行に繰り返し,
        # 金額・判定など伝票単位の値のみ先頭日の行に入れる (ソート対策 + 多重計上防止).
        "日付", "日次",
        # 3. 定時外の移動時間 勤務実態 (2026-07-17 客先依頼: 労務・健康管理の直後に組み込み.
        # 移動開始/終了・勤務開始/終了の生データはこの3.の12列でより詳細に持つため,
        # 2.側は重複する単一集計の生データ列を廃止し, 公式判定に使う3列のみ残す.
        # 2026-07-23 客先依頼: その3列 (通常経路到着時間/整合/労務実態) も削除)
        "定時前_精算開始", "定時前_精算終了", "定時前_勤怠開始", "定時前_勤怠終了",
        "勤務_精算開始", "勤務_精算終了", "勤務_勤怠開始", "勤務_勤怠終了",
        "定時後_精算開始", "定時後_精算終了", "定時後_勤怠開始", "定時後_勤怠終了",
        "定時前差分", "定時前チェック", "勤務差分", "勤務チェック", "定時後差分", "定時後チェック",
        "日当金額", "滞在費補助金額", "宿泊費（ホテル代）", "日当計", "宿泊費・手当の整合",
        "合計金額", "金額確認",
        "二重申請確認", "証跡・領収書確認",
        "総合判定", "要確認項目",
    ]),
    ("02_二次承認詳細", "secondary", [
        "伝票No.", "入力者名", "明細No.", "明細日付", "開始", "終了",
        "出発地", "到着地", "交通機関", "金額", "証票",
        "日当CD", "宿泊CD", "滞在CD", "勘定科目名",
        "照合顧客名", "距離区分", "照合状態", "複数候補",
    ]),
    ("03_差異一覧", "diff", [
        "伝票No.", "入力者名", "観点", "判定", "判定理由", "確認先システム", "対応案",
    ]),
    ("04_差戻し文面候補", "reject", [
        "伝票No.", "入力者名", "宛先(メール)", "理由区分", "判定", "差戻し文面候補",
    ]),
    ("05_取込ログ", "import_log", [
        "区分", "ファイル名", "件数", "詳細", "結果",
    ]),
    ("06_判定ルール", "rules", [
        "項目", "値", "備考",
    ]),
    ("07_マスタ確認", "master_check", [
        "種別", "対象", "詳細", "対応",
    ]),
]

# ステータス値 -> 塗りつぶし色 (ARGB). 未確認(...) 系は前方一致でグレー.
_FILL_OK = PatternFill("solid", fgColor="C6EFCE")        # 緑
_FILL_WARN = PatternFill("solid", fgColor="FFEB9C")      # 黄
_FILL_NG = PatternFill("solid", fgColor="FFC7CE")        # 赤
_FILL_GRAY = PatternFill("solid", fgColor="D9D9D9")      # 灰 (未突合/未確認)
_FILL_BANNER = PatternFill("solid", fgColor="FFF2CC")    # バナー薄黄

_STATUS_FILL = {
    OK: _FILL_OK,
    NEEDS_CHECK: _FILL_WARN,
    NG: _FILL_NG,
    MULTI: _FILL_NG,            # §6.1: 複数候補 は赤
    UNMATCHED: _FILL_GRAY,
    RULE_MISSING: _FILL_GRAY,
    ATT_MISSING: _FILL_GRAY,
}
# 文字列一致しない派生語 (checksheet が出す簡易語) もカバー
_EXTRA_FILL = {
    "突合": _FILL_OK,
    "別名突合": _FILL_WARN,
    "承認済": _FILL_OK,
    "OK": _FILL_OK,
    "不整合": _FILL_NG,
    "未承認(PENDING)": _FILL_WARN,
    "複数候補": _FILL_NG,
}

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")    # 見出し薄青
_FILL_BAND = PatternFill("solid", fgColor="F2F2F2")       # 日付グループ帯 (交互の薄灰)
_BANNER_FONT = Font(bold=True, color="9C5700")
_WRAP = Alignment(vertical="top", wrap_text=True)

def _fill_for(value) -> PatternFill | None:
    """セル値からステータス塗りを決定. 該当しなければ None.

    完全一致 (OK/要確認/NG/未突合/承認済 等) を優先し, 次に '未確認' 前方一致で灰.
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    if s in _STATUS_FILL:
        return _STATUS_FILL[s]
    if s in _EXTRA_FILL:
        return _EXTRA_FILL[s]
    if s.startswith("未確認"):
        return _FILL_GRAY
    return None


def _columns_for(rows, fixed_cols) -> list:
    """固定列 + dict に現れる未定義キーを末尾に追補して列順を確定.

    先頭が '_' のキーは内部用 (帯の grouping キー等) とみなし列にしない.
    """
    cols = list(fixed_cols)
    seen = set(cols)
    for row in rows:
        if not isinstance(row, dict):
            continue
        for k in row.keys():
            if k not in seen and not str(k).startswith("_"):
                seen.add(k)
                cols.append(k)
    return cols


def _write_data_rows(ws, rows, cols, data_start_row: int, group_col: str | None = None) -> int:
    """データ行のみを ws に書き出す (見出しは呼び出し側で書き済み). 最終行 (1-based) を返す.

    group_col 指定時: その列の値が変わるたびに帯を交互に塗り, グループを
    視覚的に見やすくする (ステータス色が付くセルは上書きしない).
    """
    r = data_start_row - 1
    band_on = False
    prev_group_val = object()  # 初回に必ず切替わるよう番兵
    for row in rows:
        r += 1
        get = row.get if isinstance(row, dict) else (lambda k, d="": d)
        if group_col is not None:
            gv = get(group_col, "")
            if gv != prev_group_val:
                band_on = not band_on
                prev_group_val = gv
        for ci, name in enumerate(cols, start=1):
            val = get(name, "")
            if val is None:
                val = ""
            cell = ws.cell(row=r, column=ci, value=val)
            fill = _fill_for(val)
            if fill is None and group_col is not None and band_on:
                fill = _FILL_BAND
            if fill is not None:
                cell.fill = fill
    return r


def _write_table(ws, rows, fixed_cols, start_row: int, group_col: str | None = None) -> int:
    """見出し (1行) + データ行を ws に書き出す. 最終行 (1-based) を返す.

    - 見出し: 太字 + 薄青塗り
    - ステータス語彙セル: 色分け
    - freeze_panes は見出し直下に設定
    """
    cols = _columns_for(rows, fixed_cols)
    # 見出し行
    for ci, name in enumerate(cols, start=1):
        c = ws.cell(row=start_row, column=ci, value=name)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _WRAP

    r = _write_data_rows(ws, rows, cols, start_row + 1, group_col=group_col)

    # 見出し直下で freeze (列見出し常時表示). 見出し行のみのときも下端を固定.
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # オートフィルタ (2026-07-31 客先要望: 各列で絞り込めるようにする).
    # 見出し行を基準に, データがある範囲へ設定する.
    if cols:
        last_col = get_column_letter(len(cols))
        ws.auto_filter.ref = f"A{start_row}:{last_col}{max(r, start_row)}"

    # 列幅 (見出しと値の最大長から概算, CJK は 2 幅換算)
    _autofit(ws, cols, rows, start_row)
    return r


def _disp_width(v) -> int:
    """表示幅概算: 全角(東アジア幅 W/F)を 2, それ以外 1 として数える."""
    import unicodedata
    s = "" if v is None else str(v)
    w = 0
    for ch in s:
        w += 2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1
    return w


def _autofit(ws, cols, rows, header_row: int) -> None:
    """列幅をヘッダ/データの最大表示幅から決定 (上限 60)."""
    for ci, name in enumerate(cols, start=1):
        width = _disp_width(name)
        for row in rows:
            if isinstance(row, dict):
                width = max(width, _disp_width(row.get(name, "")))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(width + 2, 8), 60)


def _write_banner(ws, banners) -> int:
    """サマリ先頭にバナー (既知欠落等) をマージ太字で書く. 次の空き行を返す.

    1 行目: タイトル. 以降 1 行 1 バナー. 表は banners ブロックの 2 行下から.
    """
    title = "出張精算 承認チェックシート — 既知の前提・データ欠落 (必ず確認)"
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=12, color="9C5700")
    cell.fill = _FILL_BANNER
    cell.alignment = _WRAP
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    r = 1
    for line in banners or []:
        r += 1
        c = ws.cell(row=r, column=1, value=f"・{line}")
        c.font = _BANNER_FONT
        c.fill = _FILL_BANNER
        c.alignment = _WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    # バナーと表の間に 1 行空ける
    return r + 2


def write_excel(check_sheet, out_path: str, cfg) -> None:
    """check_sheet (dict) を 7 シートの xlsx として out_path に書き出す.

    引数:
      check_sheet: build_check_sheet() の dict (上記 INPUT CONTRACT). キー欠落耐性あり.
      out_path:    出力 .xlsx パス.
      cfg:         Config (バナーのフォールバック等に使用).
    返り値: なし (ファイルを書く).
    """
    # dict 以外 (duck-typed オブジェクト) でも .get で扱えるようにする
    if hasattr(check_sheet, "get") and callable(check_sheet.get):
        cs_get = check_sheet.get
    else:
        cs_get = lambda k, d=None: getattr(check_sheet, k, d)

    wb = Workbook()
    # 既定で作られる Sheet を最初のシートとして再利用
    first = True
    for sheet_name, key, fixed_cols in SHEETS:
        rows = cs_get(key, None) or []
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        if sheet_name == "01_一次承認チェック":
            banners = cs_get("banners", None)
            if not banners:
                banners = list(getattr(cfg, "known_gaps", []) or [])
            table_start = _write_banner(ws, banners)
            # 2026-07-31: 日単位に展開したため, 伝票が切り替わるたびに帯を塗って
            # 「どこからどこまでが1伝票か」を見分けられるようにする.
            _write_table(ws, rows, fixed_cols, table_start, group_col="_group")
        elif sheet_name == "02_二次承認詳細":
            # 客先要望 (2026-07-13): 日付単位で分けて見やすくする
            _write_table(ws, rows, fixed_cols, 1, group_col="明細日付")
        else:
            _write_table(ws, rows, fixed_cols, 1)

    wb.save(out_path)
