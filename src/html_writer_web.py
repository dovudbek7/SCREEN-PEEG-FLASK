"""Excel チェックシート → 自己完結 HTML ビューワー生成.

使い方:
    from html_writer_web import write_html
    write_html(sheet_data, out_path)

sheet_data は excel_writer.py が受け取る同じ辞書構造を想定しているが、
既存の Excel ファイルからも読み込める (read_excel_to_html を参照)。
"""
from __future__ import annotations

import html
import json
import os
import re
from datetime import datetime


def _translate_cell_val(val: str, i18n: dict) -> str:
    if not i18n.get("value_map") and not i18n.get("prefix_map"):
        return val
    vm = i18n.get("value_map", {})

    # exact match
    if val in vm:
        return vm[val]

    # dot-separated compound (出張実態・労務・金額規程 etc.)
    if "・" in val and not any(c in val for c in ("(", "/")):
        parts = val.split("・")
        if all(p in vm for p in parts):
            return " • ".join(vm[p] for p in parts)

    # prefix map
    for ja_pfx, uz_pfx in i18n.get("prefix_map", []):
        if val.startswith(ja_pfx):
            return uz_pfx + val[len(ja_pfx):]

    # dynamic regex patterns
    # 日次ラベル (2026-07-31): '1日目' / '3日目(最終日)'
    m = re.fullmatch(r'(\d+)日目(\(最終日\))?', val)
    if m:
        return f"{m.group(1)}-kun" + (" (oxirgi)" if m.group(2) else "")

    # 勤務窓のはみ出しラベル (2026-08-06): '出勤前0:30' / '退勤後1:30／出勤前0:05'
    if "出勤前" in val or "退勤後" in val:
        out = val.replace("出勤前", "ishga kelishdan oldin ")
        out = out.replace("退勤後", "ishdan chiqqach ")
        return out.replace("／", " / ")

    m = re.match(r'対象件数\(伝票\)=(\d+) / 未承認件数=(\d+) / 承認済=(\d+)', val)
    if m:
        return f"Hujjat soni={m.group(1)} / Tasdiqlanmagan={m.group(2)} / Tasdiqlangan={m.group(3)}"

    m = re.match(r'勤怠日欠落 \(明細(.+)\)', val)
    if m:
        return f"Davomati yo'q (qator {m.group(1)})"

    m = re.match(r'深夜移動 要確認 \(明細(.+)\)', val)
    if m:
        return f"Kech tun harakati tekshiruvi (qator {m.group(1)})"

    m = re.match(r'非労働日/休暇に移動 \(明細(.+)\)', val)
    if m:
        return f"Dam olish kuni harakat (qator {m.group(1)})"

    m = re.match(r'金額上限超過 \((\d+)件\)', val)
    if m:
        return f"Summa limiti oshdi ({m.group(1)} ta)"

    m = re.match(r'金額上限超過 \((\d+)件\) / 要確認 \((\d+)件\)', val)
    if m:
        return f"Summa limiti oshdi ({m.group(1)} ta) / Tekshirish kerak ({m.group(2)} ta)"

    m = re.match(r'二重申請疑い: 強一致の重複明細が (\d+) 件。', val)
    if m:
        return f"Ikki marta ariza gumonasi: {m.group(1)} ta takror qator topildi"

    m = re.match(r'勤怠データ欠落: (\S+) 月分未提供', val)
    if m:
        return f"Davomati yo'q: {m.group(1)} oy taqdim etilmagan"

    # 伝票No.XXXX: 出張日が特定できません (no date range)
    m = re.match(r'伝票No\.([^:]+): 出張日\(移動/到着明細\)が特定できません。$', val)
    if m:
        return f"Hujjat No.{m.group(1)}: Safari sanasi (harakat/borish qatorlari) aniqlanmadi."

    # 伝票No.XXXX(PERIOD): MM oy davomati yo'q → borish nuqtasi topilmadi
    m = re.match(
        r'伝票No\.([^(]+)\(([^)]+)\): (.+)の出勤簿が未提供のため出張実態は暫定評価です。'
        r'到着地が顧客マスタと突合できないため、訪問先の妥当性を別途ご確認ください。$', val)
    if m:
        return (
            f"Hujjat No.{m.group(1)}({m.group(2)}): {m.group(3)} oy davomati taqdim etilmagan — "
            f"safari holati vaqtinchalik baholandi. "
            f"Borish nuqtasi mijozlar bazasida topilmadi, tashrif maqsadini alohida tekshiring."
        )

    # 伝票No.XXXX(PERIOD): MM oy davomati yo'q → davomati kelib tushgach
    m = re.match(
        r'伝票No\.([^(]+)\(([^)]+)\): (.+)の出勤簿が未提供のため出張実態は暫定評価です。'
        r'出勤簿提供後に最終確認をお願いします。$', val)
    if m:
        return (
            f"Hujjat No.{m.group(1)}({m.group(2)}): {m.group(3)} oy davomati taqdim etilmagan — "
            f"safari holati vaqtinchalik baholandi. Davomati kelib tushgach yakuniy tekshiruv."
        )

    # 伝票No.XXXX(PERIOD): DAYS ta'til/ishdan qolish kuni
    m = re.match(
        r'伝票No\.([^(]+)\(([^)]+)\): (.+)は勤怠上 休暇/欠勤 です。'
        r'出張実態と整合しないため、日付の誤りか休暇申請の取消をご確認ください。$', val)
    if m:
        return (
            f"Hujjat No.{m.group(1)}({m.group(2)}): {m.group(3)} — "
            f"davomat bo'yicha ta'til/ishdan qolish kuni. "
            f"Safari holati bilan mos kelmaydi; sana xatosi yoki ta'til bekor qilishni tekshiring."
        )

    # 伝票No.XXXX(PERIOD): DAYS davomat kiritilmagan
    m = re.match(
        r'伝票No\.([^(]+)\(([^)]+)\): (.+)の勤怠が未入力です。出勤簿の入力状況をご確認ください。$',
        val)
    if m:
        return (
            f"Hujjat No.{m.group(1)}({m.group(2)}): {m.group(3)} — "
            f"davomat kiritilmagan. Davomat varaqasi to'ldirilganligini tekshiring."
        )

    # 伝票No.XXXX(PERIOD): borish nuqtasi mijozlar bazasida topilmadi
    m = re.match(
        r'伝票No\.([^(]+)\(([^)]+)\): 到着地が顧客マスタと突合できません。'
        r'訪問先\(取引先\)名の表記をご確認ください。$', val)
    if m:
        return (
            f"Hujjat No.{m.group(1)}({m.group(2)}): "
            f"Borish nuqtasi mijozlar bazasida topilmadi. "
            f"Tashrif nomi (hamkor) yozilishini tekshiring."
        )

    # 伝票No.XXXX(PERIOD): safari holati etarli dalil yo'q
    m = re.match(
        r'伝票No\.([^(]+)\(([^)]+)\): 出張実態の裏付けが不足しています。'
        r'出勤簿および訪問先をご確認ください。$', val)
    if m:
        return (
            f"Hujjat No.{m.group(1)}({m.group(2)}): "
            f"Safari holati uchun etarli dalil yo'q. Davomat va tashrif joyini tekshiring."
        )

    # 想定承認者 NAME は確認のみ (approval_route dynamic)
    m = re.match(
        r'想定承認者 (.+) は確認のみのため正式承認に該当しません。正規の出張命令者による承認をご確認ください。$',
        val)
    if m:
        return (
            f"Tasdiqlash: {m.group(1)} faqat 'tekshirish' rolidadir — rasmiy tasdiqlash hisoblanmaydi. "
            f"To'g'ri safari buyrug'i beruvchi tomonidan tasdiqlashni tekshiring."
        )

    # 想定の出張命令者 NAME による承認が確認できません (approval_route dynamic)
    m = re.match(
        r'想定の出張命令者 (.+) による承認が確認できません。承認ルートをご確認ください。$', val)
    if m:
        return (
            f"Safari buyrug'i beruvchi {m.group(1)} tomonidan tasdiqlash topilmadi. "
            f"Tasdiqlash zanjirini tekshiring."
        )

    # compound with semicolons (non-labour + missing)
    if "; " in val:
        parts = val.split("; ")
        translated = [_translate_cell_val(p, i18n) for p in parts]
        return "; ".join(translated)

    return val


# ── ステータスラベル → CSS クラス ──────────────────────────────────────────
def _status_class(val: str) -> str:
    v = str(val).strip()
    if v == "NG":
        return "s-ng"
    if v in ("要確認", "要確認(勤怠データ欠落)"):
        return "s-warn"
    if v == "OK":
        return "s-ok"
    if v.startswith("未確認"):
        return "s-unknown"
    return ""


def _match_status_class(val: str) -> str:
    """02_二次承認詳細 の「照合状態」列 (突合/別名突合/未突合/複数候補) 用."""
    v = str(val).strip()
    if v == "突合":
        return "s-ok"
    if v == "別名突合":
        return "s-warn"
    if v == "複数候補":
        return "s-ng"
    if v == "未突合":
        return "s-unknown"
    return ""


# 01_一次承認 の列インデックス (0-based). checksheet.py primary_rows のキー挿入順と一致させる.
# 2026-07-17 客先依頼: '3. 定時外の移動時間 勤務実態' (18列) を労務・健康管理の直後に
# 組み込み. 併せて 2. 側の重複する単一集計の生データ列 (移動開始/終了・勤務開始/終了)
# を廃止し, 公式判定に使う通常経路到着時間/整合/労務実態のみ残した (3列, 旧7列).
# 2026-07-23 客先依頼: その通常経路到着時間/整合/労務実態の3列も削除 (定時外18列のみ残す).
COL01_NO = 0
COL01_VOUCHER = 1
COL01_INPUTTER = 2
COL01_EMP_NO = 3
COL01_DEPT = 4
COL01_APPROVAL_STATE = 5
COL01_PERIOD = 6
COL01_DEST_PREF = 7
COL01_CUSTOMER = 8
COL01_TRIP_REALITY = 9
COL01_DATE = 10                  # 2026-07-31 客先依頼: 日単位表示のため追加
COL01_DAYNO = 11                 # 「1日目/2日目(最終日)」ラベル
COL01_OH_BEFORE_EXP_START = 12
COL01_OH_BEFORE_EXP_END = 13
COL01_OH_BEFORE_ATT_START = 14
COL01_OH_BEFORE_ATT_END = 15
COL01_OH_WORK_EXP_START = 16
COL01_OH_WORK_EXP_END = 17
COL01_OH_WORK_ATT_START = 18
COL01_OH_WORK_ATT_END = 19
COL01_OH_AFTER_EXP_START = 20
COL01_OH_AFTER_EXP_END = 21
COL01_OH_AFTER_ATT_START = 22
COL01_OH_AFTER_ATT_END = 23
COL01_OH_BEFORE_DIFF = 24
COL01_OH_BEFORE_CHECK = 25
COL01_OH_WORK_DIFF = 26
COL01_OH_WORK_CHECK = 27
COL01_OH_AFTER_DIFF = 28
COL01_OH_AFTER_CHECK = 29
COL01_PERDIEM = 30
COL01_STAY_ALLOWANCE = 31
COL01_LODGING = 32
COL01_PERDIEM_TOTAL = 33
COL01_LODGING_MATCH = 34
COL01_TOTAL = 35
COL01_AMOUNT_RULE = 36
COL01_DUP = 37
COL01_RECEIPT = 38
COL01_OVERALL = 39
COL01_FLAGGED = 40

STATUS_COLS_01 = {
    COL01_TRIP_REALITY, COL01_AMOUNT_RULE, COL01_DUP,
    COL01_RECEIPT, COL01_OVERALL,
    COL01_OH_BEFORE_CHECK, COL01_OH_WORK_CHECK, COL01_OH_AFTER_CHECK,
}

# 中身が時刻(HH:MM)のみの列. 見出し文言に対して内容が短いので, 幅を詰めて表示する
# (客先指摘 2026-07-17: 時刻列が場所を取りすぎる).
NARROW_COLS_01 = {
    COL01_OH_BEFORE_EXP_START, COL01_OH_BEFORE_EXP_END,
    COL01_OH_BEFORE_ATT_START, COL01_OH_BEFORE_ATT_END,
    COL01_OH_WORK_EXP_START, COL01_OH_WORK_EXP_END,
    COL01_OH_WORK_ATT_START, COL01_OH_WORK_ATT_END,
    COL01_OH_AFTER_EXP_START, COL01_OH_AFTER_EXP_END,
    COL01_OH_AFTER_ATT_START, COL01_OH_AFTER_ATT_END,
}

# ステータス凡例アイコン (✓ 正常 / ▲ 要確認 / ✕ NG / － データなし)
_STATUS_ICON = {"s-ok": "✓", "s-warn": "▲", "s-ng": "✕", "s-unknown": "－"}

# 01_一次承認 の列グループ (見出し2段表示用). 各要素は (グループ名, 列スパン数).
# header 全列を左から連続でカバーする (ラベル空文字は無地スパン).
SHEET01_GROUPS = [
    ("", 5),                          # No./伝票No./入力者名/社員番号/所属
    ("承認状況", 1),                   # 承認状態
    # 2026-07-31 客先依頼: 日単位表示のため 日付 列を追加 (4列 -> 5列).
    ("1. 出張実態の確認", 6),           # 出張期間/出張先/取引先/出張実態/日付/日次
    # 2026-07-17 客先依頼: 定時外の移動時間 勤務実態 を独立区分にせず, 労務・健康管理と
    # 同じ区分にまとめる. 2026-07-23 客先依頼で通常経路到着時間/整合/労務実態の3列を
    # 削除したため, 現在は定時外18列のみ.
    ("2. 労務・健康管理の確認", 18),
    ("3. 出張費・宿泊費上限確認", 7),    # 日当金額/滞在費補助金額/宿泊費(ホテル代)/日当計/宿泊整合/合計金額/金額確認
    ("4. 全体チェック", 3),             # 二重申請確認/証跡・領収書確認/総合判定
    ("詳細", 1),                       # 要確認項目
]

# 18列グループ「2. 労務・健康管理の確認」(定時外の移動時間 勤務実態,
# COL01_OH_BEFORE_EXP_START..COL01_OH_AFTER_CHECK) を3段見出しにするための
# 追加行 (客先要望 2026-07-22: 統合前のグループ区分が一目で分かるようにして
# ほしい). 各リストは41列全体をカバーする (ラベル空文字は空白セル).
def _group_class_map(groups: list[tuple[str, int]]) -> dict[int, str]:
    """列インデックス -> 区分クラス名 (g0..gN). 区分の先頭列には gstart も付ける.

    2026-08-05 客先依頼: 5つの区分の切れ目が見やすいよう, 区分ごとに
    データの背景色を変える (CSS の .g0〜 と .gstart を参照).
    """
    mapping: dict[int, str] = {}
    col = 0
    for gi, (_label, span) in enumerate(groups):
        for k in range(span):
            cls = f"g{gi}"
            if k == 0 and gi > 0:
                cls += " gstart"
            mapping[col] = cls
            col += 1
    return mapping


SHEET01_OH_TIER_A = [
    ("", 12),
    ("定時前（9:00までの時間）", 4),
    ("勤務時間（9:00〜17:30）", 4),
    ("定時後（17:30以降の時間）", 4),
    ("勤務実態の差分・チェック", 6),
    ("", 11),
]

SHEET01_OH_TIER_B = [
    ("", 12),
    ("楽々精算", 2), ("楽々勤怠", 2),
    ("楽々精算", 2), ("楽々勤怠", 2),
    ("楽々精算", 2), ("楽々勤怠", 2),
    ("定時前移動", 2), ("勤務時間", 2), ("定時後移動", 2),
    ("", 11),
]

SHEET01_SUB_TIERS = [SHEET01_OH_TIER_A, SHEET01_OH_TIER_B]


def _cell_html(val: str, col_idx: int, sheet_id: str, i18n: dict | None = None,
               voucher_no: str = "", match_col: int = -1) -> str:
    cls = ""
    if sheet_id == "01" and col_idx in STATUS_COLS_01:
        cls = _status_class(val)
    elif sheet_id == "02" and col_idx == match_col:
        cls = _match_status_class(val)
    elif sheet_id == "03" and col_idx == 3:
        cls = _status_class(val)
    elif sheet_id == "04" and col_idx == 4:
        cls = _status_class(val)
    elif sheet_id == "05" and col_idx == 4:
        cls = "s-ok" if val == "OK" else "s-ng"
    display = _translate_cell_val(val, i18n) if i18n else val
    if cls:
        icon = _STATUS_ICON.get(cls, "")
        # 01の要確認/NG/未確認等 (OK以外) は 03_差異一覧 へジャンプ可能にする
        # (客先要望 2026-07-13: 要確認からリンクで詳細へ跳べるように).
        if sheet_id == "01" and col_idx in STATUS_COLS_01 and cls != "s-ok" and voucher_no:
            v_esc = html.escape(voucher_no, quote=True)
            hint = (i18n or {}).get("jump_hint", "")
            return (f'<span class="badge {cls} jump-badge" tabindex="0" '
                    f'onclick="jumpToDiff(\'{v_esc}\', event)" '
                    f'onkeypress="if(event.key===\'Enter\')jumpToDiff(\'{v_esc}\', event)" '
                    f'title="{html.escape(hint, quote=True)}">{icon} {display}</span>')
        return f'<span class="badge {cls}">{icon} {display}</span>'
    return display


def _spans_to_map(spans: list[tuple[str, int]]) -> dict[int, tuple[str, int]]:
    """[(label, span), ...] を 開始列インデックス -> (label, span) の辞書にする."""
    m: dict[int, tuple[str, int]] = {}
    pos = 0
    for label, span in spans:
        m[pos] = (label, span)
        pos += span
    return m


def _header_rows_html(header: list[str], groups: list, tr: dict, narrow_cols: set | None = None,
                      tr_full: dict | None = None,
                      sub_tiers: list[list[tuple[str, int]]] | None = None) -> str:
    """2段見出し (グループ行 + 列名行) を組み立てる. sub_tiers 指定時は
    _header_rows_html_with_sub_tiers に委譲し, 4段見出しを組み立てる
    (客先要望 2026-07-22: 中間段の空白行が3段分積み重なって縦に間延びして
    見えるのを解消するため, 非対象列は rowspan で1セルに結合し中央寄せする).
    """
    narrow_cols = narrow_cols or set()
    tr_full = tr_full or {}
    total = sum(span for _, span in groups)
    if total != len(header):
        # 列数不一致なら安全に通常の1段見出しへフォールバック
        th_html = "".join(
            f'<th onclick="sortTable(this)" data-col="{i}">'
            f'{tr.get(h, h)}<span class="sort-icon">⇅</span></th>'
            for i, h in enumerate(header)
        )
        return f"<tr>{th_html}</tr>"

    if sub_tiers and len(sub_tiers) >= 2 and all(
        sum(span for _, span in tier) == len(header) for tier in sub_tiers
    ):
        return _header_rows_html_with_sub_tiers(header, groups, tr, narrow_cols, tr_full, sub_tiers)

    # ラベル無し(無地)グループの列見出しは, 1段目に空セル・2段目に列名セルを
    # 1列ずつ並べる (見た目はrowspan=2と同じ2段ぶち抜きになる). rowspanは
    # 意図的に使わない — th に position:sticky (JSで付与, スクロール固定用) と
    # rowspan を同時に使うと, 行(tr)自体が非stickyのまま子セルだけ視覚的に
    # はがれて別位置に描画され, テーブルの行順序が崩れて見える不具合が
    # Chrome等で発生するため (2026-07-15 客先報告で発覚, スクリーンショット確認済み).
    row1: list[str] = []
    row2: list[str] = []
    start = 0
    for label, span in groups:
        if label:
            row1.append(
                f'<th colspan="{span}" class="grp-th" data-start="{start}" '
                f'data-span="{span}" onclick="highlightGroup(this)" '
                f'onmouseenter="previewGroup(this,true)" onmouseleave="previewGroup(this,false)">'
                f'{tr.get(label, label)}</th>'
            )
            for i in range(start, start + span):
                h = header[i]
                cls = ' class="col-narrow"' if i in narrow_cols else ""
                full = tr_full.get(h) or (tr.get(h, h) if i in narrow_cols else "")
                full_attr = f' data-full="{html.escape(full, quote=True)}"' if full else ""
                row2.append(
                    f'<th onclick="sortTable(this)" data-col="{i}"{cls}{full_attr}>'
                    f'{tr.get(h, h)}<span class="sort-icon">⇅</span></th>'
                )
        else:
            for i in range(start, start + span):
                row1.append('<th class="grp-solo-blank"></th>')
            for i in range(start, start + span):
                h = header[i]
                row2.append(
                    f'<th onclick="sortTable(this)" data-col="{i}" class="grp-solo">'
                    f'{tr.get(h, h)}<span class="sort-icon">⇅</span></th>'
                )
        start += span

    return f"<tr class='grp-row'>{''.join(row1)}</tr><tr>{''.join(row2)}</tr>"


def _header_rows_html_with_sub_tiers(header: list[str], groups: list, tr: dict,
                                     narrow_cols: set, tr_full: dict,
                                     sub_tiers: list[list[tuple[str, int]]]) -> str:
    """4段見出し (グループ行 + 中間段2行 + 列名行) を組み立てる.

    中間段 (sub_tiers) に実ラベルが無い列 (定時外ブロック以外の全列) は,
    3行 (中間段2行+列名行) または4行 (グループ行含む, 無地グループの場合) を
    rowspan で1セルに結合し, 縦横センタリングする (客先要望 2026-07-22:
    空白行が3段積み重なって縦に間延びして見える問題の解消). rowspanは
    定時外ブロック以外でのみ使う — 現状 thead th に position:sticky は
    どこにも付与していないため (grep確認済み), _header_rows_html の
    docstringが警告するChromeの不具合 (rowspan + JS付与sticky) は起きない.
    将来スクロール追従ヘッダー等でsticky を再導入する場合は, その対象が
    ここで rowspan を使うセルと重ならないか確認すること.
    """
    tier_a, tier_b = sub_tiers[0], sub_tiers[1]
    n = len(header)
    tier_a_starts = _spans_to_map(tier_a)
    tier_b_starts = _spans_to_map(tier_b)

    offhours_cols: set[int] = set()
    for pos, (label, span) in tier_a_starts.items():
        if label:
            offhours_cols.update(range(pos, pos + span))

    is_blank_col = [False] * n
    pos = 0
    for label, span in groups:
        if not label:
            for j in range(pos, pos + span):
                is_blank_col[j] = True
        pos += span

    gcls = _group_class_map(groups)

    def leaf_th(i: int, rowspan: int = 1) -> str:
        h = header[i]
        cls_parts = []
        if i in gcls:
            cls_parts.append(gcls[i])
        if i in narrow_cols:
            cls_parts.append("col-narrow")
        if rowspan > 1:
            cls_parts.append("grp-merged")
        cls_attr = f' class="{" ".join(cls_parts)}"' if cls_parts else ""
        full = tr_full.get(h) or (tr.get(h, h) if i in narrow_cols else "")
        full_attr = f' data-full="{html.escape(full, quote=True)}"' if full else ""
        rs_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ""
        return (f'<th onclick="sortTable(this)" data-col="{i}"{rs_attr}{cls_attr}{full_attr}>'
                f'{tr.get(h, h)}<span class="sort-icon">⇅</span></th>')

    # 1行目: 業務グループ (既存 SHEET01_GROUPS のロジックそのまま). 無地グループ
    # (定時外ブロックとは重ならない前提, 5列固定) だけは各列を rowspan=4 の
    # 実列名セルに置き換える (旧: 空セル+2段目に実列名).
    row1: list[str] = []
    start = 0
    for label, span in groups:
        if label:
            row1.append(
                f'<th colspan="{span}" class="grp-th" data-start="{start}" '
                f'data-span="{span}" onclick="highlightGroup(this)" '
                f'onmouseenter="previewGroup(this,true)" onmouseleave="previewGroup(this,false)">'
                f'{tr.get(label, label)}</th>'
            )
        else:
            for i in range(start, start + span):
                row1.append(leaf_th(i, rowspan=4))
        start += span

    # 2-4行目 (中間段A/B + 列名): 定時外ブロック(offhours_cols)は実ラベルを
    # そのまま各行に配置. それ以外は無地グループ(rowspan=4で1行目に既出のため
    # ここではスキップ)か, ラベル付きグループの列(rowspan=3で中間段Aに結合).
    tier_a_cells: list[str] = []
    tier_b_cells: list[str] = []
    leaf: list[str] = []
    i = 0
    while i < n:
        if i in offhours_cols:
            a_label, a_span = tier_a_starts[i]
            tier_a_cells.append(f'<th colspan="{a_span}" class="grp-mid grp-tier-a">{tr.get(a_label, a_label)}</th>')
            j = i
            while j < i + a_span:
                b_label, b_span = tier_b_starts[j]
                tier_b_cells.append(f'<th colspan="{b_span}" class="grp-mid grp-tier-b">{tr.get(b_label, b_label)}</th>')
                j += b_span
            for j in range(i, i + a_span):
                leaf.append(leaf_th(j))
            i += a_span
        elif is_blank_col[i]:
            i += 1
        else:
            tier_a_cells.append(leaf_th(i, rowspan=3))
            i += 1

    return (f"<tr class='grp-row'>{''.join(row1)}</tr>"
            f"<tr class='grp-row-mid grp-tier-a'>{''.join(tier_a_cells)}</tr>"
            f"<tr class='grp-row-mid grp-tier-b'>{''.join(tier_b_cells)}</tr>"
            f"<tr>{''.join(leaf)}</tr>")


def _build_table(header: list[str], rows: list[list[str]], sheet_id: str,
                 col_translations: dict | None = None, i18n: dict | None = None,
                 groups: list | None = None,
                 sub_tiers: list[list[tuple[str, int]]] | None = None) -> str:
    tr = col_translations or {}
    tr_full = (i18n or {}).get("headers_full", {})
    narrow_cols = NARROW_COLS_01 if sheet_id == "01" else set()
    if groups:
        header_html = _header_rows_html(header, groups, tr, narrow_cols, tr_full, sub_tiers)
    else:
        th_html = "".join(
            f'<th onclick="sortTable(this)" data-col="{i}">'
            f'{tr.get(h, h)}<span class="sort-icon">⇅</span></th>'
            for i, h in enumerate(header)
        )
        header_html = f"<tr>{th_html}</tr>"
    date_col = header.index("明細日付") if "明細日付" in header else -1
    match_col = header.index("照合状態") if sheet_id == "02" and "照合状態" in header else -1
    # 単一ステータス列を持つ他シート (客先要望 2026-07-13: フィルタを全シートへ展開)
    simple_status_col = {"02": match_col, "03": 3, "04": 4, "05": 4}.get(sheet_id, -1)
    # 列インデックス -> 区分クラス (背景色の塗り分け用). groups が無いシートは空。
    body_gcls = _group_class_map(groups) if groups else {}
    # 01シートの日単位展開でグループ (伝票) を追跡するための可変ホルダ
    _sheet01_group = [""]
    tr_html_parts = []
    for row in rows:
        voucher_no = ""
        is_head = day_of_group = False
        if sheet_id == "01":
            if len(row) > COL01_VOUCHER:
                voucher_no = str(row[COL01_VOUCHER])
            is_head = bool(len(row) > COL01_OVERALL and str(row[COL01_OVERALL]).strip())
            day_of_group = not is_head
        cells = []
        for i in range(len(header)):
            raw_val = str(row[i]) if i < len(row) else ""
            translated = _translate_cell_val(raw_val, i18n) if i18n else raw_val
            full_attr = f' data-full="{html.escape(translated)}"' if translated else ""
            cls_parts = []
            if i in body_gcls:
                cls_parts.append(body_gcls[i])
            if i in narrow_cols:
                cls_parts.append("col-narrow")
            cls_attr = f' class="{" ".join(cls_parts)}"' if cls_parts else ""
            inner = _cell_html(raw_val, i, sheet_id, i18n, voucher_no, match_col)
            cells.append(f'<td{full_attr}{cls_attr}>{inner}</td>')
        tds = "".join(cells)
        extra_attrs = ""
        if sheet_id == "01":
            # 2026-07-31: 01シートは 伝票 が日数分の行に展開されている。
            # 伝票No. 等の識別列は全行に繰り返されるため, グループ先頭 (伝票単位の値を
            # 持つ行) の判定には 総合判定 が入っているかを使う。
            # フィルタ/検索/ソートをグループ単位で効かせるため, 全行に同じ data-grp を振る。
            grp_id = str(row[COL01_VOUCHER]).strip() if len(row) > COL01_VOUCHER else ""
            if is_head:
                _sheet01_group[0] = grp_id
                extra_attrs += ' data-grphead="1"'
            elif not grp_id:
                grp_id = _sheet01_group[0]
            extra_attrs += f' data-grp="{html.escape(grp_id, quote=True)}"'
            if day_of_group:
                # 2日目以降の行 (識別列は淡色表示にして先頭日と見分けやすくする).
                # 2026-08-04 客先依頼: 折りたたみ (プルダウン) は廃止し, 全行を常時表示する.
                extra_attrs += ' class="grp-day"'
            if len(row) > COL01_OVERALL:
                raw_status = str(row[COL01_OVERALL])
                code = "未確認" if raw_status.startswith("未確認") else raw_status
                extra_attrs += f' data-status="{code}"'
            # 各観点 (総合判定以外) も個別フィルタ可能にする
            for ax_col in STATUS_COLS_01:
                if ax_col == COL01_OVERALL or ax_col >= len(row):
                    continue
                raw_ax = str(row[ax_col])
                code = "未確認" if raw_ax.startswith("未確認") else raw_ax
                extra_attrs += f' data-ax{ax_col}="{html.escape(code, quote=True)}"'
        elif simple_status_col >= 0 and simple_status_col < len(row):
            raw_status = str(row[simple_status_col])
            code = "未確認" if raw_status.startswith("未確認") else raw_status
            extra_attrs += f' data-status="{html.escape(code, quote=True)}"'
        tr_html_parts.append(f"<tr{extra_attrs}>{tds}</tr>")
    tbody = "\n".join(tr_html_parts)
    return f"""
<div class="tbl-wrap">
  <table class="data-tbl" id="tbl-{sheet_id}" data-date-col="{date_col}">
    <thead>{header_html}</thead>
    <tbody>{tbody}</tbody>
  </table>
</div>"""


I18N = {
    "ja": {
        "title":        "出張精算 承認チェックシート ビューワー",
        "generated":    "生成日時",
        "search":       "🔍 検索...",
        "count_suffix": "件",
        "lbl_ng":       "NG",
        "lbl_warn":     "要確認",
        "lbl_ok":       "OK",
        "lbl_unknown":  "未確認",
        "lbl_total":    "合計件数",
        "legend_title": "ステータス凡例：",
        "filter_status_all": "ステータス：すべて",
        "compact_on":  "表を広く表示",
        "compact_off": "件数・フィルタを表示",
        "compact_hint": "バナー・件数サマリー・絞り込みを隠して, 表を画面いっぱいに表示します",
        "jump_hint": "クリックで詳細（03_差異一覧）へ移動",
        "filter_bar_title": "観点別フィルタ：",
        "sort_by_date": "日付で絞り込み",
        "date_placeholder": "選択または入力...",
        "tabs": {
            "01": "01_一次承認",
            "02": "02_二次明細",
            "03": "03_差異一覧",
            "04": "04_差戻し文面",
            "05": "05_取込ログ",
            "06": "06_判定ルール",
            "07": "07_マスタ確認",
        },
        "lang_attr": "ja",
        "font": '"Hiragino Kaku Gothic ProN","Yu Gothic",Meiryo,"Noto Sans JP",sans-serif',
    },
    "uz": {
        "title":        "Xizmat safari xarajatlari tasdiqlash varag'i",
        "generated":    "Yaratilgan sana",
        "search":       "🔍 Qidirish...",
        "count_suffix": "ta",
        "lbl_ng":       "NG (Rad)",
        "lbl_warn":     "Tekshirish kerak",
        "lbl_ok":       "OK",
        "lbl_unknown":  "Noaniq",
        "lbl_total":    "Jami ariza",
        "legend_title": "Status belgilari:",
        "filter_status_all": "Status: barchasi",
        "compact_on":  "Jadvalni kengaytirish",
        "compact_off": "Statistika va filtrni ko'rsatish",
        "compact_hint": "Banner, statistika va filtrlarni yashirib, jadvalni to'liq ekranga chiqaradi",
        "jump_hint": "Bosib tafsilotga (03_Farqlar) o'ting",
        "filter_bar_title": "Mezon bo'yicha filter:",
        "sort_by_date": "Sana bo'yicha filter",
        "date_placeholder": "Tanlang yoki yozing...",
        "tabs": {
            "01": "01 — Birlamchi tasdiqlash",
            "02": "02 — Tafsilot",
            "03": "03 — Farqlar ro'yxati",
            "04": "04 — Qaytarish matni",
            "05": "05 — Yuklash jurnali",
            "06": "06 — Qoidalar",
            "07": "07 — Master tekshiruvi",
        },
        "lang_attr": "uz",
        "font": '"Inter","Segoe UI",system-ui,sans-serif',
        "prefix_map": [
            ("取込日時 ", "Yuklash vaqti "),
        ],
        "value_map": {
            # ── 承認状態 ──
            "承認済":                         "Tasdiqlangan",
            "承認中":                         "Tasdiqlash jarayonida",
            "未承認":                         "Tasdiqlanmagan",
            "未承認(PENDING)":                "Tasdiqlanmagan",
            "未確認(勤怠データ欠落)":         "Noaniq (davomati yo'q)",
            # ── Sheet 01 '3.定時外の移動時間 勤務実態' の差分・チェック値 ──
            "一致":                           "Mos keladi",
            "データなし":                     "Ma'lumot yo'q",
            "移動なし":                       "Harakat yo'q",
            "勤務時間内":                     "Ish vaqti ichida",
            "勤怠打刻なし":                   "Davomatda belgi yo'q",
            "精算計上なし":                   "Arizada yo'q",
            "作業計上なし":                   "Ish qatori yo'q",
            "未確認(突合不可)":               "Aniqlanmagan (solishtirib bo'lmadi)",
            # ── 判定 ──
            "要確認":                         "Tekshirish kerak",
            "要確認(勤怠データ欠落)":         "Tekshirish kerak (davomati yo'q)",
            # ── Sheet 02 証票 ──
            "証票あり":                       "Kvitansiya bor",
            "証票なし":                       "Kvitansiya yo'q",
            "免除":                           "Mustasnо",
            # ── Sheet 02 照合状態 ──
            "突合":                           "Mos keldi",
            "別名突合":                       "Taxminiy mos",
            "未突合":                         "Mos kelmadi",
            # ── Sheet 03 観点 (compound parts) ──
            "出張実態":                       "Safari holati",
            "労務":                           "Mehnat",
            "金額規程":                       "Summa qoidasi",
            "二重申請":                       "Ikki marta ariza",
            "領収書":                         "Kvitansiya",
            "承認ルート":                     "Tasdiqlash zanjiri",
            # ── 判定理由 (static known strings) ──
            "旅費規定の上限を超過した明細の妥当性を確認してください。":
                "Safar xarajatlari qoidasidan oshgan qatorlarni tekshiring",
            "金額がマイナス/小計不一致の明細を是正してください。":
                "Manfiy summa yoki yig'indi mos kelmaydigan qatorlarni to'g'rilang",
            "深夜帯の移動について時刻・深夜手当の要否をご確認ください。":
                "Kech tun harakati vaqti va tungi to'lov zaruratini tekshiring",
            "承認者ルート判定不可: 申請者がマスタ/名簿に未登録":
                "Tasdiqlash zanjiri aniqlanmadi: ariza beruvchi ro'yxatda yo'q",
            "免除交通機関だが高額の明細に領収書がありません。領収書の添付を依頼してください。":
                "Mustasnо transport bo'lsa-da yuqori summali qatorda kvitansiya yo'q. Kvitansiya so'rang",
            # ── Transport turlari ──
            "電車･ﾊﾞｽ":                      "Poyezd/Avtobus",
            "飛行機":                         "Samolyot",
            "車(同乗)":                       "Mashina (birgalikda)",
            "ﾀｸｼｰ(同乗)":                   "Taksi (birgalikda)",
            "レンタカー":                     "Ijaraga mashina",
            "テレワーク":                     "Masofaviy ish",
            "徒歩":                           "Piyoda",
            "作業･打合せ":                    "Ish/Yig'ilish",
            "ホテル":                         "Mehmonxona",
            # ── Xarajat kategoriyalari ──
            "ホテル代":                       "Mehmonxona to'lovi",
            "宿泊税":                         "Mehmonxona solig'i",
            "入湯税":                         "Hammom solig'i",
            "ガソリン代":                     "Benzin",
            "ｶﾞｿﾘﾝ代":                       "Benzin",
            "コインパーキング":               "Pulli avtoturargoh",
            "駐車代":                         "Avtoturargoh to'lovi",
            "駐車代金":                       "Avtoturargoh to'lovi",
            "駐車場":                         "Avtoturargoh",
            "駐車場代":                       "Avtoturargoh to'lovi",
            "旅費交通費":                     "Sayohat/transport xarajatlari",
            "委託サービス費":                 "Shartnoma xizmat haqi",
            "賃借料":                         "Ijara to'lovi",
            # ── Tekshirish tizimlari ──
            "楽々勤怠 (勤務実績)":            "Davomati tizimi (ish natijalari)",
            "楽々勤怠 (勤務時刻/休日)":       "Davomati tizimi (vaqt/dam olish)",
            "楽々精算 (他申請)":              "Xarajat tizimi (boshqa ariza)",
            "楽々精算 (添付/証票)":           "Xarajat tizimi (ilova/kvitansiya)",
            "楽々精算 + 旅費規定":            "Xarajat tizimi + safari qoidalari",
            "20期承認者名簿 + 楽々精算":      "20-davr tasdiqlovchilar ro'yxati + xarajat tizimi",
            # ── Boshqa ──
            "(集計)":                         "(Yig'indi)",
            "複数候補":                       "Ko'p nomzod",
            # ── Sheet 05 区分 ──
            "出張精算CSV":                    "Xizmat safari CSV",
            "勤怠(出勤簿)":                   "Davomati (ish vaqti kitobi)",
            "社員マスタ":                     "Xodimlar ro'yxati",
            "顧客マスタ":                     "Mijozlar ro'yxati",
            "承認者名簿":                     "Tasdiqlovchilar ro'yxati",
            "件数照合":                       "Sonlar tekshiruvi",
            # ── Sheet 05 結果 ──
            "エラー":                         "Xato",
            # ── Sheet 06 項目 ──
            "氏名ファジー閾値":               "Ism taxminiy chegarasi",
            "地名照合閾値":                   "Joylashuv moslashtirish chegarasi",
            "深夜発 閾値":                    "Kech tun jo'nab ketish chegarasi",
            "深夜着 閾値":                    "Kech tun yetib kelish chegarasi",
            "領収書 高額暫定閾値":            "Kvitansiya yuqori qiymat chegarasi",
            "領収書 検知下限":                "Kvitansiya minimal tekshiruv summasi",
            "金額規程(上限)提供":             "Summa qoidasi (yuqori limit) mavjud",
            "領収書要否規程提供":             "Kvitansiya talab qoidasi mavjud",
            "確認のみ=承認 扱い":             "Faqat 'tekshirish' = tasdiqlash sanaladi",
            "出張日当_一般職":                "Safari kunlik to'lov — oddiy xodim",
            "出張日当_管理職":                "Safari kunlik to'lov — boshqaruv",
            "滞在補助費_一般職":              "Qolish yordami — oddiy xodim",
            "滞在補助費_管理職":              "Qolish yordami — boshqaruv",
            "出張加算日当_一般職":            "Safari qo'shimcha kunlik — oddiy xodim",
            "出張加算日当_主任以上":          "Safari qo'shimcha kunlik — katta mutaxassis+",
            "ホテル代_東京23区_管理職":       "Mehmonxona (Tokio 23 tuman) — boshqaruv",
            "ホテル代_東京23区_一般職":       "Mehmonxona (Tokio 23 tuman) — oddiy xodim",
            "ホテル代_その他_管理職":         "Mehmonxona (boshqa hududlar) — boshqaruv",
            "ホテル代_その他_一般職":         "Mehmonxona (boshqa hududlar) — oddiy xodim",
            "既知の前提/欠落":                "Ma'lum ogohlantirishlar",
            # ── Sheet 06 値 ──
            "あり":                           "mavjud",
            "しない":                         "yo'q",
            # ── Sheet 06 備考 ──
            "difflib ratio (社員突合)":        "difflib ratio (xodim moslashtirish)",
            "rapidfuzz partial_ratio (顧客突合)": "rapidfuzz partial_ratio (mijoz moslashtirish)",
            "移動開始がこれ以前":              "Harakat boshlash vaqti bundan oldin bo'lsa",
            "移動終了がこれ以降":              "Harakat tugash vaqti bundan keyin bo'lsa",
            "規程未提供時の高額判定":          "Qoida taqdim etilmaganda yuqori summa tekshiruvi",
            "これ未満の小額は要確認にしない":  "Bundan kam kichik summalar tekshirilmaydi",
            "未提供時は金額上限を要確認(NGは出さない)":
                "Taqdim etilmaganda summa limiti tekshiruv (NG chiqarilmaydi)",
            "未提供時は暫定閾値で判定":        "Taqdim etilmaganda vaqtinchalik chegara ishlatiladi",
            "(確認)承認者を正式承認と見なすか":
                "(Tekshirish) tasdiqlovchi rasmiy tasdiqlash sanalsinmi",
            "円/日":                          "¥/kun",
            "円":                             "¥",
            # known_gaps strings (sheet 06 備考 column)
            "2026-05 の出勤簿(勤怠)が未提供 — 出張実態・労務の勤怠照合は劣化(advisory)。":
                "2026-05 oy davomati taqdim etilmagan — safari holati va mehnat tekshiruvi sifati pasaygan (advisory)",
            "社員マスタに役職列(役職/グレード/職位)が未登録の場合、役職不明として一般職上限と管理職上限の間は要確認扱いになります。":
                "Xodim ro'yxatida lavozim ustuni bo'lmasa, lavozim noaniq sifatida oddiy xodim va boshqaruv limiti orasidagi summa tekshiruv talab qiladi",
            # ── Sheet 07 種別 ──
            "承認者名簿 未登録":              "Tasdiqlovchilar ro'yxatida yo'q",
            "名簿/マスタ不整合":              "Ro'yxat/Master nomuvofiqlik",
            # ── Sheet 07 詳細 ──
            "申請者が20期承認者名簿に未登録 (出張命令者を判定不可)":
                "Ariza beruvchi 20-davr tasdiqlovchilar ro'yxatida yo'q (safari buyrug'i beruvchini aniqlab bo'lmaydi)",
            "20期名簿に存在するが社員マスタに無い氏名":
                "20-davr ro'yxatida bor, lekin xodimlar ro'yxatida yo'q ism",
            # ── Sheet 07 対応 ──
            "20期名簿に申請者を登録":
                "20-davr ro'yxatiga ariza beruvchini qo'shing",
            "社員マスタ/名簿の氏名表記を突合・統一":
                "Xodimlar ro'yxati/nomlar ro'yxatidagi ism yozuvini solishtiring va birlashtiring",
            # ── 対応案 / 差戻し文面 (static suggestion strings) ──
            "承認手続きが未完了です。承認実行のうえ再提出してください。":
                "Tasdiqlash jarayoni tugallanmagan. Tasdiqlang va qayta yuboring.",
            "申請者を従業員マスタ/承認者名簿(20期)に登録のうえ再判定してください。":
                "Ariza beruvchini xodimlar/tasdiqlovchilar ro'yxatiga (20-davr) qo'shing va qayta tekshiring.",
            "非労働日/休暇日の移動について出張の必要性・実態をご確認ください。":
                "Dam olish/ta'til kunidagi harakat uchun safari zaruratini tekshiring.",
            "該当月の勤怠データが未提供のため労務照合は劣化しています。勤怠提供後に再確認してください。":
                "Tegishli oy davomati taqdim etilmagan — mehnat tekshiruvi sifati pasaygan. Davomati kelib tushgach qayta tekshiring.",
            "労務 (勤怠整合) について手動でご確認ください。":
                "Mehnat (davomat muvofiqligi) ni qo'lda tekshiring.",
            "同一日・同一区間・同額の明細が他伝票と重複していないか確認し、必要なら一方を取消。":
                "Bir xil sana/marshrut/summa bilan boshqa hujjatda takror qator bor-yo'qligini tekshiring; kerak bo'lsa birini bekor qiling.",
            "同日・同一顧客先への複数申請の妥当性を確認。":
                "Bir kunda bir xil mijozga bir nechta ariza asosliligini tekshiring.",
            "明細金額の合計が申告合計を超過しています。金額を確認・修正してください。":
                "Qatorlar summasi e'lon qilingan umumiy summadan oshib ketgan. Miqdorni tekshiring.",
            "申告合計と明細合計の差額(手当コードなし)の内訳を確認してください。":
                "E'lon qilingan va qatorlar summasi farqi (to'lovsiz) sababini tekshiring.",
            "行数(宣言)と明細件数の差異を確認してください。":
                "E'lon qilingan qatorlar soni bilan haqiqiy qatorlar soni mos kelmaydi.",
            "金額がマイナス/小計不一致の明細を是正してください。":
                "Manfiy summa yoki yig'indi mos kelmaydigan qatorlarni to'g'rilang.",
            "旅費規定の上限を超過した明細の妥当性を確認してください。":
                "Safar xarajatlari qoidasidan oshgan qatorlarni tekshiring.",
        },
        "notice_map": {
            "2026-05 の出勤簿(勤怠)が未提供":
                "・2026-05 oy davomati (ish vaqti) taqdim etilmagan — safari holati va mehnat tekshiruvi sifati pasaygan (advisory)",
            "社員マスタに役職列":
                "・Xodim ro'yxatida lavozim ustuni (lavozim/daraja/mansab) ro'yxatga olinmagan bo'lsa, lavozim noaniq sifatida oddiy xodim va boshqaruv limiti orasidagi summa tekshiruv talab qiladi",
            "勤怠データ対象月":
                "・Davomati hisobot oylari",
        },
        "headers": {
            # ── Sheet 01 列グループ見出し ──
            "承認状況":                       "Tasdiqlash holati",
            "1. 出張実態の確認":              "1. Safar holati tekshiruvi",
            "2. 労務・健康管理の確認":        "2. Mehnat/sog'liq tekshiruvi",
            "3. 出張費・宿泊費上限確認":      "3. Xarajat/mehmonxona limiti",
            "4. 全体チェック":                "4. Umumiy tekshiruv",
            "詳細":                           "Tafsilot",
            # ── Sheet 01 ──
            "No.":          "№",
            "伝票No.":      "Hujjat №",
            "入力者名":     "Murojaat qiluvchi",
            "社員番号":     "Xodim raqami",
            "所属":         "Bo'lim",
            "出張期間":     "Safari muddati",
            "合計金額":     "Jami summa",
            "承認状態":     "Tasdiqlash holati",
            "出張先":       "Yo'nalish (hudud)",
            "取引先":       "Mijoz/kompaniya",
            "出張実態":     "Safari holati",
            "日付":         "Sana",
            "日次":         "Kun",
            "移動開始時間": "Harakat boshlanishi",
            "移動終了時間": "Harakat tugashi",
            "勤務開始時間": "Ish boshlanishi",
            "勤務終了時間": "Ish tugashi",
            "日当金額":     "Kunlik to'lov",
            "滞在費補助金額": "Qolish yordam puli",
            "宿泊費（ホテル代）": "Mehmonxona to'lovi",
            "日当計":       "Kunlik to'lov jami",
            "宿泊費・手当の整合": "Mehmonxona/to'lov moslik",
            "金額確認":     "Summa tekshiruvi",
            "二重申請確認": "Ikki marta ariza tekshiruvi",
            "証跡・領収書確認": "Kvitansiya tekshiruvi",
            "総合判定":     "Umumiy baholash",
            "要確認項目":   "Tekshirilishi kerak",
            # ── Sheet 02 ──
            "明細No.":      "Qator №",
            "明細日付":     "Sana",
            "開始":         "Boshlanish",
            "終了":         "Tugash",
            "出発地":       "Jo'nab ketish joyi",
            "到着地":       "Borish joyi",
            "交通機関":     "Transport",
            "金額":         "Summa",
            "証票":         "Kvitansiya holati",
            "日当CD":       "Kunlik to'lov kodi",
            "宿泊CD":       "Yotoqxona kodi",
            "滞在CD":       "Qolish kodi",
            "勘定科目名":   "Hisobvaraq nomi",
            "照合顧客名":   "Moslashtirish mijozi",
            "距離区分":     "Masofa turi",
            "照合状態":     "Moslashtirish holati",
            "複数候補":     "Ko'p nomzod",
            # ── Sheet 03 ──
            "観点":         "Tekshiruv nuqtai nazari",
            "判定":         "Baholash",
            "判定理由":     "Baholash sababi",
            "確認先システム": "Tekshirish tizimi",
            "対応案":       "Tavsiya",
            # ── Sheet 04 ──
            "宛先(メール)": "Qabul qiluvchi (email)",
            "理由区分":     "Sabab turi",
            "差戻し文面候補": "Qaytarish matni",
            # ── Sheet 05 ──
            "区分":         "Tur",
            "ファイル名":   "Fayl nomi",
            "件数":         "Soni",
            "詳細":         "Tafsilot",
            "結果":         "Natija",
            # ── Sheet 06 ──
            "項目":         "Parametr",
            "値":           "Qiymat",
            "備考":         "Izoh",
            # ── Sheet 07 ──
            "種別":         "Tur",
            "対象":         "Ob'ekt",
            "対応":         "Chora",
            # ── Sheet 01 '3. 定時外の移動時間 勤務実態' (労務・健康管理の直後に組み込み,
            #    2026-07-17客先依頼) ──
            "定時前_精算開始": "Seisan bosh.(O)", "定時前_精算終了": "Seisan tug.(O)",
            "定時前_勤怠開始": "Kintai bosh.(O)", "定時前_勤怠終了": "Kintai tug.(O)",
            "勤務_精算開始": "Seisan bosh.(I)", "勤務_精算終了": "Seisan tug.(I)",
            "勤務_勤怠開始": "Kintai bosh.(I)", "勤務_勤怠終了": "Kintai tug.(I)",
            "定時後_精算開始": "Seisan bosh.(K)", "定時後_精算終了": "Seisan tug.(K)",
            "定時後_勤怠開始": "Kintai bosh.(K)", "定時後_勤怠終了": "Kintai tug.(K)",
            "定時前差分": "Farq (oldin)", "定時前チェック": "Tekshiruv (oldin)",
            "勤務差分": "Farq (ish)", "勤務チェック": "Tekshiruv (ish)",
            "定時後差分": "Farq (keyin)", "定時後チェック": "Tekshiruv (keyin)",
            # ── Sheet 01 定時外3段見出し (グループ/中間段ラベル, 2026-07-22客先依頼) ──
            "定時前（9:00までの時間）": "Ish boshlanishidan oldin (09:00 gacha)",
            "勤務時間（9:00〜17:30）":   "Ish vaqti (09:00–17:30)",
            "定時後（17:30以降の時間）": "Ish tugagandan keyin (17:30 dan keyin)",
            "勤務実態の差分・チェック":  "Farq va tekshiruv (ish faoliyati)",
            "楽々精算":     "Xarajat tizimi (Rakuraku Seisan)",
            "楽々勤怠":     "Davomat tizimi (Rakuraku Kintai)",
            "定時前移動":   "Harakat (ish boshlanishidan oldin)",
            "勤務時間":     "Ish vaqti",
            "定時後移動":   "Harakat (ish tugagandan keyin)",
        },
        # 列見出しが省略表示(...)される狭い列 (NARROW_COLS_01) の hover ツールチップ用
        # フル表記. headers辞書の短縮版とは別に保持する (客先要望 2026-07-17).
        "headers_full": {
            "定時前_精算開始": "Seisan boshlanishi (定時前/oldin)",
            "定時前_精算終了": "Seisan tugashi (定時前/oldin)",
            "定時前_勤怠開始": "Kintai boshlanishi (定時前/oldin)",
            "定時前_勤怠終了": "Kintai tugashi (定時前/oldin)",
            "勤務_精算開始": "Seisan boshlanishi (勤務時間/ish)",
            "勤務_精算終了": "Seisan tugashi (勤務時間/ish)",
            "勤務_勤怠開始": "Kintai boshlanishi (勤務時間/ish)",
            "勤務_勤怠終了": "Kintai tugashi (勤務時間/ish)",
            "定時後_精算開始": "Seisan boshlanishi (定時後/keyin)",
            "定時後_精算終了": "Seisan tugashi (定時後/keyin)",
            "定時後_勤怠開始": "Kintai boshlanishi (定時後/keyin)",
            "定時後_勤怠終了": "Kintai tugashi (定時後/keyin)",
        },
    },
}


def _sheet01_stats(rows: list[list[str]], i18n: dict) -> str:
    # 2026-07-31: 01シートは日単位に展開されており, 2日目以降の行は総合判定が空欄。
    # 件数は「伝票が何件か」を示すべきなので, 総合判定が入っている行だけを数える。
    c = COL01_OVERALL
    ng      = sum(1 for r in rows if len(r) > c and r[c] == "NG")
    warn    = sum(1 for r in rows if len(r) > c and r[c] == "要確認")
    ok      = sum(1 for r in rows if len(r) > c and r[c] == "OK")
    unknown = sum(1 for r in rows if len(r) > c and r[c].startswith("未確認"))
    total   = sum(1 for r in rows if len(r) > c and r[c].strip())
    suffix = i18n["count_suffix"]
    return f"""
<div class="stats-panel">
  <div class="legend-card">
    <div class="legend-card-title">{i18n["legend_title"]}</div>
    <div class="legend-card-item"><span class="lg-ico ic-ok">✓</span>{i18n["lbl_ok"]}</div>
    <div class="legend-card-item"><span class="lg-ico ic-warn">▲</span>{i18n["lbl_warn"]}</div>
    <div class="legend-card-item"><span class="lg-ico ic-ng">✕</span>{i18n["lbl_ng"]}</div>
    <div class="legend-card-item"><span class="lg-ico ic-unk">－</span>{i18n["lbl_unknown"]}</div>
  </div>
  <table class="stat-table">
    <thead>
      <tr>
        <th class="st-ng">{i18n["lbl_ng"]}</th>
        <th class="st-warn">{i18n["lbl_warn"]}</th>
        <th class="st-ok">{i18n["lbl_ok"]}</th>
        <th class="st-unk">{i18n["lbl_unknown"]}</th>
        <th class="st-total">{i18n["lbl_total"]}</th>
      </tr>
    </thead>
    <tbody>
      <tr>
        <td class="st-ng">{ng}{suffix}</td>
        <td class="st-warn">{warn}{suffix}</td>
        <td class="st-ok">{ok}{suffix}</td>
        <td class="st-unk">{unknown}{suffix}</td>
        <td class="st-total">{total}{suffix}</td>
      </tr>
    </tbody>
  </table>
</div>"""


def _render_sheet(sid: str, label: str, data: dict, i18n: dict) -> str:
    header = data["header"]
    rows = data["rows"]
    notices_html = ""
    if "notices" in data and data["notices"]:
        notice_map = i18n.get("notice_map", {})
        def _translate_notice(n: str) -> str:
            for ja_key, uz_val in notice_map.items():
                if ja_key in n:
                    if "対象月" in ja_key:
                        suffix = n.split(":")[-1].strip() if ":" in n else ""
                        return uz_val + (f": {suffix}" if suffix else "")
                    return uz_val
            return n
        items = "".join(f"<li>{_translate_notice(n)}</li>" for n in data["notices"])
        notices_html = f'<div class="notice-box"><ul>{items}</ul></div>'
    stats_html = _sheet01_stats(rows, i18n) if sid == "01" else ""
    suffix = i18n["count_suffix"]
    if sid == "01":
        status_opts = "".join(
            f'<option value="{v}">{lbl}</option>'
            for v, lbl in [
                ("OK", i18n["lbl_ok"]), ("要確認", i18n["lbl_warn"]),
                ("NG", i18n["lbl_ng"]), ("未確認", i18n["lbl_unknown"]),
            ]
        )
        # 総合判定以外の各観点 (出張実態/労務/金額/二重申請/領収書, および
        # 定時外の移動時間 勤務実態の3チェック列) も個別に絞り込めるフィルタチップを
        # 並べる (客先要望 2026-07-13: 全項目でソート/絞込).
        tr_headers = i18n.get("headers", {})
        # 2026-07-31: 01シートは日単位に展開されているため, 初期表示の件数も
        # 行数ではなく伝票件数にする (統計パネルの合計件数と一致させる).
        voucher_count = sum(
            1 for r in rows
            if len(r) > COL01_OVERALL and str(r[COL01_OVERALL]).strip()
        )
        axis_cols = sorted(c for c in STATUS_COLS_01 if c != COL01_OVERALL)
        chip_label = lambda c: tr_headers.get(header[c], header[c]) if c < len(header) else ""
        overall_html = f"""
    <div class="filter-chip">
      <label for="status-{sid}">{i18n['legend_title']}</label>
      <select class="status-select" id="status-{sid}" onchange="applyFiltersMulti('{sid}', '{suffix}')">
        <option value="">{i18n['filter_status_all']}</option>
        {status_opts}
      </select>
    </div>"""
        chip_html = "".join(
            f'<div class="filter-chip">'
            f'<label for="filter-{sid}-{c}">{chip_label(c)}</label>'
            f'<select id="filter-{sid}-{c}" onchange="applyFiltersMulti(\'{sid}\', \'{suffix}\')">'
            f'<option value="">{i18n["filter_status_all"]}</option>{status_opts}</select>'
            f'</div>'
            for c in axis_cols if c < len(header)
        )
        search_html = f"""
<div class="toolbar toolbar-01">
  <div class="toolbar-row">
    <input class="search-box" type="text" placeholder="{i18n['search']}" oninput="applyFiltersMulti('{sid}', '{suffix}')" id="search-{sid}">
    <span class="row-count" id="count-{sid}">{voucher_count} {suffix}</span>
    <button type="button" class="compact-btn" onclick="toggleCompact(this)"
            data-on="{i18n.get('compact_on', '')}" data-off="{i18n.get('compact_off', '')}"
            title="{i18n.get('compact_hint', '')}">{i18n.get('compact_on', '')}</button>
  </div>
  <div class="filter-bar">
    <span class="filter-bar-title">{i18n.get('filter_bar_title', '')}</span>
    {overall_html}
    {chip_html}
  </div>
</div>"""
    else:
        # 単一ステータス列を持つシートにも検索+ステータス絞込を展開する
        # (客先要望 2026-07-13: 01のフィルタを全シートへ).
        vm = i18n.get("value_map", {})
        status_vocab = {
            "02": ["突合", "別名突合", "未突合", "複数候補"],
            "03": ["OK", "要確認", "NG", "未確認"],
            "04": ["OK", "要確認", "NG", "未確認"],
            "05": ["OK", "エラー"],
        }.get(sid)
        if status_vocab:
            opts = "".join(f'<option value="{v}">{vm.get(v, v)}</option>' for v in status_vocab)
            date_picker = ""
            if sid == "02" and "明細日付" in header:
                dc = header.index("明細日付")
                unique_dates = sorted({
                    str(row[dc]) for row in rows if dc < len(row) and row[dc]
                })
                date_opts = "".join(f'<option value="{d}"></option>' for d in unique_dates)
                date_picker = f"""
    <div class="filter-chip">
      <label for="datepick-{sid}">{i18n.get('sort_by_date', '')}</label>
      <input class="date-picker" id="datepick-{sid}" list="datelist-{sid}"
             placeholder="{i18n.get('date_placeholder', '')}"
             oninput="applyFilterGeneric('{sid}', '{suffix}')" autocomplete="off">
      <datalist id="datelist-{sid}">{date_opts}</datalist>
    </div>"""
            search_html = f"""
<div class="toolbar">
  <div class="toolbar-row">
    <input class="search-box" type="text" placeholder="{i18n['search']}" oninput="applyFilterGeneric('{sid}', '{suffix}')" id="search-{sid}">
    <div class="filter-chip">
      <label for="status-{sid}">{i18n['legend_title']}</label>
      <select class="status-select" id="status-{sid}" onchange="applyFilterGeneric('{sid}', '{suffix}')">
        <option value="">{i18n['filter_status_all']}</option>
        {opts}
      </select>
    </div>
    {date_picker}
    <span class="row-count" id="count-{sid}">{len(rows)} {suffix}</span>
  </div>
</div>"""
        else:
            search_html = f"""
<div class="toolbar">
  <input class="search-box" type="text" placeholder="{i18n['search']}" oninput="filterTable(this, '{sid}', '{suffix}')" id="search-{sid}">
  <span class="row-count" id="count-{sid}">{len(rows)} {suffix}</span>
</div>"""
    groups = SHEET01_GROUPS if sid == "01" else None
    sub_tiers = SHEET01_SUB_TIERS if sid == "01" else None
    table_html = _build_table(header, rows, sid, i18n.get("headers"), i18n, groups=groups, sub_tiers=sub_tiers)
    return f"""
<div id="panel-{sid}" class="panel" style="display:none" data-suffix="{suffix}">
  {notices_html}
  {stats_html}
  {search_html}
  {table_html}
</div>"""


CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
:root {
  --navy: #1b3a6b;
  --navy2: #14305a;
  --accent: #2563eb;
  --ok: #166534;
  --ok-bg: #dcfce7;
  --ng: #991b1b;
  --ng-bg: #fee2e2;
  --warn: #92400e;
  --warn-bg: #fef3c7;
  --unk: #374151;
  --unk-bg: #f3f4f6;
  --border: #d1d5db;
  --stripe: #f8fafc;
  --text: #111827;
  --radius: 6px;
}
body {
  font-family: "Hiragino Kaku Gothic ProN","Yu Gothic",Meiryo,"Noto Sans JP",sans-serif;
  font-size: 12.5px;
  color: var(--text);
  background: #f1f5f9;
  /* 2026-08-06: 見出し固定のためページ自体はスクロールさせず, 表の中だけを
     スクロールさせる. ページと表の二重スクロールになると, 上部のツールバーが
     ヘッダーの下に潜り込んで見切れてしまうため. */
  height: 100vh;
  overflow: hidden;
  display: flex;
  flex-direction: column;
}
/* ── Header ── */
.app-header {
  background: var(--navy);
  color: #fff;
  padding: 14px 24px 0;
  flex: 0 0 auto;
  z-index: 100;
  box-shadow: 0 2px 8px rgba(0,0,0,0.25);
}
.app-title { font-size: 15px; font-weight: 700; letter-spacing: 0.04em; margin-bottom: 10px; }
.app-meta { font-size: 11px; color: rgba(255,255,255,0.65); margin-bottom: 10px; }
/* ── Tabs ── */
.tabs { display: flex; gap: 2px; }
.tab-btn {
  padding: 8px 14px;
  background: rgba(255,255,255,0.12);
  color: rgba(255,255,255,0.75);
  border: none;
  border-radius: 6px 6px 0 0;
  cursor: pointer;
  font-size: 12px;
  font-family: inherit;
  white-space: nowrap;
  transition: background 0.15s;
}
.tab-btn:hover { background: rgba(255,255,255,0.22); color:#fff; }
.tab-btn.active { background: #fff; color: var(--navy); font-weight: 700; }
/* ── Main content ── */
.main {
  padding: 10px 20px;
  flex: 1 1 auto;
  min-height: 0;          /* flex 子が縮めるようにする */
  display: flex;
  flex-direction: column;
}
.panel {
  animation: fadeIn 0.15s ease;
  flex: 1 1 auto;
  min-height: 0;
  display: flex;
  flex-direction: column;
}
@keyframes fadeIn { from { opacity:0; transform:translateY(4px); } to { opacity:1; transform:none; } }
/* ── Notice box ── */
.notice-box {
  background: #fff8e1;
  border-left: 3px solid #f59e0b;
  border-radius: 0 var(--radius) var(--radius) 0;
  padding: 6px 12px;
  margin-bottom: 8px;
  font-size: 11.5px;
  color: #78350f;
}
.notice-box ul { padding-left: 16px; }
.notice-box li { margin-bottom: 3px; }
/* ── Legend + Stats (side-by-side card layout) ── */
.stats-panel { display: flex; align-items: center; gap: 12px; margin-bottom: 8px; flex-wrap: wrap; }
.legend-card {
  border: 1px solid var(--border); border-radius: var(--radius);
  background: #fff; padding: 6px 12px; font-size: 11.5px; color: var(--text);
  display: flex; align-items: center; gap: 10px; flex-wrap: wrap;
}
.legend-card-title { font-weight: 700; font-size: 11px; }
.legend-card-item { display: flex; align-items: center; gap: 4px; white-space: nowrap; }
.lg-ico { font-weight: 700; width: 14px; text-align: center; display: inline-block; }
.lg-ico.ic-ok   { color: var(--ok); }
.lg-ico.ic-warn { color: var(--warn); }
.lg-ico.ic-ng   { color: var(--ng); }
.lg-ico.ic-unk  { color: var(--unk); }
.stat-table {
  border-collapse: collapse; font-size: 12px; background: #fff;
  border-radius: var(--radius); overflow: hidden;
  box-shadow: 0 1px 4px rgba(0,0,0,0.1);
}
.stat-table th, .stat-table td {
  border: 1px solid var(--border); padding: 4px 16px; text-align: center; white-space: nowrap;
}
.stat-table thead th { font-weight: 700; background: #f8fafc; }
.stat-table tbody td { font-size: 16px; font-weight: 700; }
.stat-table .st-ng   { color: var(--ng); }
.stat-table .st-warn { color: var(--warn); }
.stat-table .st-ok   { color: var(--ok); }
.stat-table .st-unk  { color: var(--unk); }
.stat-table .st-total{ color: var(--accent); }
/* ── Toolbar ── */
.toolbar { display: flex; align-items: center; gap: 12px; margin-bottom: 6px; }
/* 2026-08-06: 表を広く使うため, 上部 (バナー/件数/フィルタ) を畳めるようにする */
.compact-btn {
  margin-left: auto; padding: 5px 12px; font-size: 11.5px; font-weight: 600;
  border: 1px solid var(--border); border-radius: var(--radius);
  background: #fff; color: var(--muted); cursor: pointer; white-space: nowrap;
  font-family: inherit;
}
.compact-btn:hover { background: #eef2ff; color: var(--accent); border-color: var(--accent); }
.panel.compact .notice-box,
.panel.compact .stats-panel,
.panel.compact .filter-bar { display: none; }
.search-box {
  padding: 7px 12px;
  border: 1px solid var(--border);
  border-radius: var(--radius);
  font-size: 12.5px;
  font-family: inherit;
  width: 320px;
  background: #fff;
}
.search-box:focus { outline: none; border-color: var(--accent); box-shadow: 0 0 0 3px rgba(37,99,235,0.15); }
.status-select {
  padding: 7px 10px;
  border: 1px solid var(--border);
  border-radius: var(--radius);
  font-size: 12.5px;
  font-family: inherit;
  background: #fff;
}
.status-select:focus { outline: none; border-color: var(--accent); box-shadow: 0 0 0 3px rgba(37,99,235,0.15); }
.row-count { font-size: 11px; color: #6b7280; white-space: nowrap; }
/* ── Toolbar (sheet 01): 検索行 + 観点別フィルタバー ── */
.toolbar-01 { display: block; margin-bottom: 6px; }
.toolbar-row { display: flex; align-items: center; gap: 12px; margin-bottom: 6px; }
.filter-bar {
  display: flex; flex-wrap: wrap; align-items: center; gap: 4px 10px;
  padding: 6px 10px; background: #f8fafc; border: 1px solid var(--border);
  border-radius: var(--radius);
}
.filter-bar-title { font-size: 11px; font-weight: 700; color: #6b7280; margin-right: 2px; white-space: nowrap; }
.filter-chip { display: flex; flex-direction: column; gap: 2px; }
.filter-chip label { font-size: 10px; color: #6b7280; font-weight: 600; white-space: nowrap; }
.filter-chip select {
  padding: 5px 8px; border: 1px solid var(--border); border-radius: 999px;
  font-size: 11.5px; font-family: inherit; background: #fff; color: var(--text);
  min-width: 120px;
}
.filter-chip select:focus { outline: none; border-color: var(--accent); box-shadow: 0 0 0 3px rgba(37,99,235,0.15); }
.date-picker {
  padding: 5px 8px; border: 1px solid var(--border); border-radius: 999px;
  font-size: 11.5px; font-family: inherit; background: #fff; color: var(--text);
  min-width: 140px;
}
.date-picker:focus { outline: none; border-color: var(--accent); box-shadow: 0 0 0 3px rgba(37,99,235,0.15); }
/* ── Sheet 01: 要確認/NG バッジは 03_差異一覧 へのリンク ── */
.jump-badge { cursor: pointer; text-decoration: none; border: 1px solid transparent; }
.jump-badge:hover, .jump-badge:focus { text-decoration: underline; border-color: currentColor; outline: none; }
/* ── Sheet 02: 明細日付ごとの帯 ── */
tr.date-band td {
  background: var(--navy); color: #fff; font-weight: 700; font-size: 11.5px;
  padding: 6px 12px; position: sticky; top: 0; z-index: 1;
}
/* ── Table ── */
.tbl-wrap {
  /* 2026-08-06 客先依頼: ヘッダーを固定する.
     表の中だけをスクロールさせ, thead を sticky で貼り付ける.
     高さは画面から上部のツールバー分を引いた値. */
  flex: 1 1 auto;
  min-height: 0;
  overflow: auto;
  border-radius: var(--radius);
  box-shadow: 0 1px 4px rgba(0,0,0,0.1);
  background: #fff;
}
/* 見出し行の固定. 01シートは4段見出しのため, 各段に上からの位置を指定する.
   段の高さを固定して top を積み上げる (可変にすると段がずれるため). */
/* sticky にすると th は tr の背景を引き継がないため, 背景は th 自身に指定する
   (指定しないと固定中の見出しが透けて, 下の行が透けて見えてしまう). */
.data-tbl thead th {
  position: sticky;
  z-index: 5;
  background: var(--navy);
}
.data-tbl thead tr.grp-row th { height: 30px; top: 0; background: #0e2144; }
.data-tbl thead tr.grp-row-mid.grp-tier-a th { height: 24px; top: 30px; background: #16345f; }
.data-tbl thead tr.grp-row-mid.grp-tier-b th { height: 24px; top: 54px; background: var(--navy); }
.data-tbl thead tr:last-child th { height: 34px; }
/* 4段見出し (01シート) の列名行 */
.data-tbl thead tr.grp-row ~ tr:last-child th { top: 78px; }
/* 1段見出し (02〜07シート) の列名行 */
.data-tbl thead tr:first-child:last-child th { top: 0; }
/* rowspan で結合された見出しは最上段から始まるので 0 に戻す */
.data-tbl thead th.grp-merged { top: 0; z-index: 6; }
.data-tbl { width: 100%; border-collapse: collapse; font-size: 12px; }
.data-tbl thead tr { background: var(--navy); color: #fff; }
/* 見出しをドラッグした際のブラウザ既定の選択ハイライト(薄灰青)を消す
   (客先報告 2026-07-17: hoverすると薄い色で残る不具合 — user-select:none だけでは
   一部ブラウザで選択ハイライトの描画が残ることがあるための保険) */
.data-tbl thead th::selection { background: transparent; }
.data-tbl thead th *::selection { background: transparent; }
.data-tbl thead th {
  padding: 9px 12px;
  text-align: left;
  white-space: nowrap;
  cursor: pointer;
  user-select: none;
  font-weight: 600;
  letter-spacing: 0.02em;
}
.data-tbl thead th:hover { background: var(--navy2); }
.sort-icon { margin-left: 4px; opacity: 0.5; font-size: 10px; }
/* ── Grouped header (sheet 01) ── */
.data-tbl thead tr.grp-row { background: #0e2144; }
.data-tbl thead th.grp-th {
  height: 38px;
  padding: 10px 16px;
  text-align: center;
  font-size: 10.5px;
  font-weight: 700;
  letter-spacing: 0.03em;
  border-left: 1px solid rgba(255,255,255,0.2);
  border-bottom: 2px solid rgba(255,255,255,0.35);
  cursor: pointer;
  white-space: nowrap;
  transition: background 0.12s;
}
.data-tbl thead th.grp-th:hover { background: #1e4a8a; }
.data-tbl thead th.grp-th.grp-active { background: var(--accent); }
.data-tbl thead th[rowspan="2"] { vertical-align: middle; }
.data-tbl thead th.grp-solo:hover { background: #1e4a8a; }
/* ── Grouped header sub-tiers (sheet 01: 定時外の移動時間 勤務実態) ── */
.data-tbl thead th.grp-mid {
  padding: 7px 10px;
  text-align: center;
  font-size: 10px;
  font-weight: 600;
  letter-spacing: 0.02em;
  border-left: 1px solid rgba(255,255,255,0.15);
  border-bottom: 1px solid rgba(255,255,255,0.25);
  white-space: nowrap;
}
.data-tbl thead tr.grp-row-mid.grp-tier-a { background: #16345f; }
.data-tbl thead tr.grp-row-mid.grp-tier-b { background: var(--navy); }
/* 定時外ブロック以外の列: 中間段2行+列名行 (または無地グループは1行目も) を
   rowspan で1セルに結合したもの. 縦横センタリングし, 背景は列名行と同じ
   navy で統一 (客先要望 2026-07-22). */
.data-tbl thead th.grp-merged {
  background: var(--navy);
  text-align: center;
  vertical-align: middle;
}
.data-tbl tbody tr:nth-child(even) { background: var(--stripe); }
.data-tbl tbody tr:hover { background: #eff6ff; }
/* ── 区分ごとの塗り分け (客先要望 2026-08-05: 5区分の切れ目を見やすく) ──
   半透明 (rgba) にしているのは, 行の縞模様とホバーを潰さないため.
   区分の先頭列には左に縦罫を入れて境界をはっきりさせる. */
.data-tbl tbody td.g0 { background: transparent; }                 /* No.〜所属 (区分外) */
.data-tbl tbody td.g1 { background: rgba(100,116,139,.13); }       /* 承認状況 */
.data-tbl tbody td.g2 { background: rgba(59,130,246,.16); }        /* 1. 出張実態の確認 */
.data-tbl tbody td.g3 { background: rgba(16,185,129,.16); }        /* 2. 労務・健康管理の確認 */
.data-tbl tbody td.g4 { background: rgba(245,158,11,.20); }        /* 3. 出張費・宿泊費上限確認 */
.data-tbl tbody td.g5 { background: rgba(167,139,250,.20); }       /* 4. 全体チェック */
.data-tbl tbody td.g6 { background: rgba(100,116,139,.13); }       /* 詳細 */
.data-tbl tbody td.gstart { border-left: 2px solid rgba(27,47,110,.3); }
.data-tbl thead th.gstart { border-left: 2px solid rgba(255,255,255,.5); }
/* 見出し側も同じ色味を薄く乗せて, どの列がどの区分かひと目で分かるようにする */
.data-tbl thead tr:last-child th.g1 { background: #2b3f7a; }
.data-tbl thead tr:last-child th.g2 { background: #24407f; }
.data-tbl thead tr:last-child th.g3 { background: #17456a; }
.data-tbl thead tr:last-child th.g4 { background: #4a3f6b; }
.data-tbl thead tr:last-child th.g5 { background: #3a2f6e; }
.data-tbl thead tr:last-child th.g6 { background: #2b3f7a; }
/* 列ハイライトとホバーは区分色より優先させる (この順序が重要) */
.data-tbl tbody td.col-hl { background: #dbeafe; }
.data-tbl tbody tr:hover td.col-hl { background: #bfdbfe; }
.data-tbl tbody td.col-hl-preview { background: #eff6ff; }
.data-tbl tbody tr:hover td.col-hl-preview { background: #dbeafe; }
/* ── 行の選択 (客先要望 2026-08-05: クリックした行を固定表示して見失わないように) ──
   区分色・ホバー・列ハイライトより後に置いて, 選択状態が最優先で見えるようにする. */
.data-tbl tbody tr.row-selected td,
.data-tbl tbody tr.row-selected:hover td,
.data-tbl tbody tr.row-selected td.col-hl,
.data-tbl tbody tr.row-selected:hover td.col-hl {
  background: #fef3c7;
}
.data-tbl tbody tr.row-selected td:first-child {
  box-shadow: inset 3px 0 0 0 #d97706;
}
.data-tbl tbody tr.row-selected td { cursor: default; }
.data-tbl tbody td {
  padding: 6px 12px;
  border-bottom: 1px solid #e5e7eb;
  vertical-align: middle;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
  max-width: 220px;
}
/* ── 時刻(HH:MM)のみの列は狭く詰める (客先指摘 2026-07-17).
   見出し文言が幅を超える場合は隣の列と重ならないよう省略表示にする. ── */
.data-tbl thead th.col-narrow, .data-tbl tbody td.col-narrow {
  padding-left: 8px;
  padding-right: 8px;
  max-width: 110px;
  overflow: hidden;
  text-overflow: ellipsis;
}
.data-tbl tbody tr.hidden { display: none; }
/* ── 複数日の伝票: 2日目以降の行 ──
   2026-08-04 客先依頼で常時表示に変更。
   2026-08-05 客先依頼: 2日目以降も文字を淡色にせず, 1日目と同じ濃さにする。
   代わりに伝票の先頭行の上に太い罫線を引き, どこから次の伝票かを示す。 */
.data-tbl tbody tr[data-grphead="1"] td { border-top: 2px solid rgba(27,47,110,.28); }
.data-tbl tbody tr:first-child td { border-top: none; }
.data-tbl tbody td[data-full] { cursor: help; }
.data-tbl thead th[data-full] { cursor: help; }
/* ── Badges ── */
.badge { display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; white-space: nowrap; }
.s-ng      { background: var(--ng-bg);   color: var(--ng);   }
.s-warn    { background: var(--warn-bg); color: var(--warn); }
.s-ok      { background: var(--ok-bg);   color: var(--ok);   }
.s-unknown { background: var(--unk-bg);  color: var(--unk);  }
/* ── Cell tooltip (hover preview / click to pin) ── */
.cell-tip {
  position: fixed; z-index: 900; display: none;
  max-width: 440px; background: #1f2937; color: #fff;
  font-size: 12.5px; line-height: 1.65; padding: 11px 15px;
  border-radius: 8px; box-shadow: 0 10px 28px rgba(0,0,0,0.3);
  white-space: normal; word-break: break-word; pointer-events: none;
}
.cell-tip.pinned {
  pointer-events: auto; border: 1px solid rgba(255,255,255,0.25);
}
.cell-tip .tip-close {
  display: none; float: right; margin: -4px -6px 0 10px;
  cursor: pointer; opacity: 0.7; font-size: 13px;
}
.cell-tip.pinned .tip-close { display: inline; }
.cell-tip.pinned .tip-close:hover { opacity: 1; }
"""

JS = """
// ── Tab switching ──
function showTab(id) {
  document.querySelectorAll('.panel').forEach(p => p.style.display = 'none');
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById('panel-' + id).style.display = 'flex';
  document.querySelector('[data-tab="' + id + '"]').classList.add('active');
}

// ── suffix (件/ta) を panel の data-suffix から解決 ──
function _resolveSuffix(sid, suffix) {
  if (suffix) return suffix;
  const panel = document.getElementById('panel-' + sid);
  return (panel && panel.dataset.suffix) || '件';
}

// ── Search / filter ──
function filterTable(input, sid, suffix) {
  suffix = _resolveSuffix(sid, suffix);
  const q = input.value.toLowerCase();
  const tbody = document.querySelector('#tbl-' + sid + ' tbody');
  let visible = 0;
  tbody.querySelectorAll('tr:not(.date-band)').forEach(tr => {
    const match = tr.textContent.toLowerCase().includes(q);
    tr.classList.toggle('hidden', !match);
    if (match) visible++;
  });
  document.getElementById('count-' + sid).textContent = visible + ' ' + suffix;
  regroupDates(sid);
}

// ── Sheet 01/08: text search + (任意)単一ステータス + 観点別ドロップダウン 全て組み合わせて絞り込む ──
function applyFiltersMulti(sid, suffix) {
  suffix = _resolveSuffix(sid, suffix);
  const searchEl = document.getElementById('search-' + sid);
  const q = searchEl ? searchEl.value.toLowerCase() : '';
  const statusEl = document.getElementById('status-' + sid);
  const status = statusEl ? statusEl.value : '';
  const axisSelects = Array.from(
    document.querySelectorAll('#panel-' + sid + ' .filter-bar select[id^="filter-' + sid + '-"]')
  );
  const tbody = document.querySelector('#tbl-' + sid + ' tbody');
  // 2026-07-31: 01シートは1伝票が日数分の行に展開されている。ステータス等の
  // 伝票単位の値は先頭行にしか無いため, フィルタは「伝票グループ」単位で判定する。
  //  - ステータス/観点フィルタ … グループ先頭行の値で判定
  //  - フリーワード検索      … グループ内のどの行に含まれていてもヒット扱い
  // 件数はグループ数 (=伝票件数) を表示し, 統計パネルと一致させる。
  const groups = new Map();
  tbody.querySelectorAll('tr:not(.date-band)').forEach(tr => {
    const key = tr.dataset.grp || ('#' + groups.size);
    if (!groups.has(key)) groups.set(key, {head: null, rows: []});
    const g = groups.get(key);
    g.rows.push(tr);
    if (tr.dataset.grphead === '1' || g.head === null) g.head = tr;
  });
  let visible = 0;
  groups.forEach(g => {
    const head = g.head;
    const textMatch = !q || g.rows.some(tr => tr.textContent.toLowerCase().includes(q));
    const statusMatch = !status || head.dataset.status === status;
    let axisMatch = true;
    axisSelects.forEach(sel => {
      if (!sel.value) return;
      const colIdx = sel.id.split('-')[2];
      if ((head.dataset['ax' + colIdx] || '') !== sel.value) axisMatch = false;
    });
    const match = textMatch && statusMatch && axisMatch;
    g.rows.forEach(tr => tr.classList.toggle('hidden', !match));
    if (match) visible++;
  });
  document.getElementById('count-' + sid).textContent = visible + ' ' + suffix;
  regroupDates(sid);
}

// ── 01以外で単一ステータス列を持つシート (02/03/04/05): 検索+ステータス絞込 ──
// (客先要望 2026-07-13: 01のフィルタを全シートへ展開)
function applyFilterGeneric(sid, suffix) {
  suffix = _resolveSuffix(sid, suffix);
  const searchEl = document.getElementById('search-' + sid);
  const statusEl = document.getElementById('status-' + sid);
  const dateEl = document.getElementById('datepick-' + sid);
  const q = searchEl ? searchEl.value.toLowerCase() : '';
  const status = statusEl ? statusEl.value : '';
  const dateQ = dateEl ? dateEl.value.trim() : '';
  const table = document.getElementById('tbl-' + sid);
  const dateColIdx = table ? parseInt(table.dataset.dateCol, 10) : -1;
  const tbody = table ? table.querySelector('tbody') : null;
  if (!tbody) return;
  let visible = 0;
  tbody.querySelectorAll('tr:not(.date-band)').forEach(tr => {
    const textMatch = tr.textContent.toLowerCase().includes(q);
    const statusMatch = !status || tr.dataset.status === status;
    let dateMatch = true;
    if (dateQ && dateColIdx >= 0) {
      const cell = tr.cells[dateColIdx];
      dateMatch = (cell ? cell.textContent.trim() : '').includes(dateQ);
    }
    const match = textMatch && statusMatch && dateMatch;
    tr.classList.toggle('hidden', !match);
    if (match) visible++;
  });
  document.getElementById('count-' + sid).textContent = visible + ' ' + suffix;
  regroupDates(sid);
}

// ── 明細日付列があるシート (02_二次明細) の行を日付ごとに帯分けする ──
// (客先要望 2026-07-13: 二次承認者のシートを日付単位で分けて見やすくする)
function regroupDates(sid) {
  const table = document.getElementById('tbl-' + sid);
  if (!table) return;
  const colIdx = parseInt(table.dataset.dateCol, 10);
  if (isNaN(colIdx) || colIdx < 0) return;
  const tbody = table.querySelector('tbody');
  tbody.querySelectorAll('tr.date-band').forEach(tr => tr.remove());
  const headRow = table.querySelector('thead tr:last-child');
  const ncols = headRow ? headRow.children.length : 1;
  let lastDate = null;
  tbody.querySelectorAll('tr').forEach(tr => {
    if (tr.classList.contains('hidden')) return;
    const cell = tr.cells[colIdx];
    const val = cell ? cell.textContent.trim() : '';
    if (val !== lastDate) {
      const band = document.createElement('tr');
      band.className = 'date-band';
      const td = document.createElement('td');
      td.colSpan = ncols;
      td.textContent = val || '—';
      band.appendChild(td);
      tr.parentNode.insertBefore(band, tr);
      lastDate = val;
    }
  });
}

// ── 01の要確認/NG等バッジをクリックすると 03_差異一覧 へ移動し, 伝票No.で絞り込む ──
// (客先要望 2026-07-13: 要確認がリンクに跳ぶように)
function jumpToDiff(voucherNo, evt) {
  if (evt) evt.stopPropagation();
  showTab('03');
  const inp = document.getElementById('search-03');
  const statusEl = document.getElementById('status-03');
  if (statusEl) statusEl.value = '';
  if (inp) {
    inp.value = voucherNo;
    applyFilterGeneric('03');
  }
  const tbl = document.getElementById('tbl-03');
  if (tbl) tbl.scrollIntoView({ behavior: 'smooth', block: 'start' });
}

// ── Group header hover: temporary column preview (removed on mouseleave) ──
function previewGroup(th, on) {
  const table = th.closest('table');
  const start = parseInt(th.dataset.start);
  const span = parseInt(th.dataset.span);
  table.querySelectorAll('tbody tr').forEach(tr => {
    for (let i = start; i < start + span; i++) {
      if (tr.cells[i]) tr.cells[i].classList.toggle('col-hl-preview', on);
    }
  });
}

// ── Group header click: highlight its columns ──
function highlightGroup(th) {
  const table = th.closest('table');
  const wasActive = th.classList.contains('grp-active');
  table.querySelectorAll('.grp-th').forEach(t => t.classList.remove('grp-active'));
  table.querySelectorAll('td.col-hl').forEach(td => td.classList.remove('col-hl'));
  if (wasActive) return;
  th.classList.add('grp-active');
  const start = parseInt(th.dataset.start);
  const span = parseInt(th.dataset.span);
  table.querySelectorAll('tbody tr').forEach(tr => {
    for (let i = start; i < start + span; i++) {
      if (tr.cells[i]) tr.cells[i].classList.add('col-hl');
    }
  });
}

// ── Sort ──
let _sortState = {};
function sortTable(th) {
  const table = th.closest('table');
  const sid = table.id.replace('tbl-', '');
  const col = parseInt(th.dataset.col);
  const key = sid + '-' + col;
  const asc = !_sortState[key];
  _sortState[key] = asc;

  // reset icons
  th.closest('thead').querySelectorAll('.sort-icon').forEach(s => s.textContent = '⇅');
  th.querySelector('.sort-icon').textContent = asc ? '↑' : '↓';

  const tbody = table.querySelector('tbody');
  const rows = Array.from(tbody.querySelectorAll('tr:not(.date-band)'));

  const cmp = (av, bv) => {
    // 空欄は常に末尾へ (昇順/降順どちらでも) — 空で並びが乱れないように
    if (av === '' && bv === '') return 0;
    if (av === '') return 1;
    if (bv === '') return -1;
    const an = parseFloat(av.replace(/[^0-9.-]/g, ''));
    const bn = parseFloat(bv.replace(/[^0-9.-]/g, ''));
    if (!isNaN(an) && !isNaN(bn)) return asc ? an - bn : bn - an;
    return asc ? av.localeCompare(bv, 'ja') : bv.localeCompare(av, 'ja');
  };
  const cellOf = (tr) => tr.cells[col] ? tr.cells[col].textContent.trim() : '';

  tbody.querySelectorAll('tr.date-band').forEach(tr => tr.remove());

  if (sid === '01') {
    // 2026-07-31 客先指摘: 01シートは1伝票が日数分の行に展開されているため,
    // 行単位でソートすると同じ伝票の行がバラバラになる。伝票グループを1つの
    // かたまりとして並べ替え, グループ内は日付順のまま保つ。
    // グループのソートキーは「先頭行(伝票単位の値を持つ行)の値」。ただし日付など
    // 日単位の列でソートした場合は, グループ内で最初に値が入っている行を使う。
    const groups = [];
    const byKey = new Map();
    rows.forEach(tr => {
      const k = tr.dataset.grp || ('#' + groups.length);
      let g = byKey.get(k);
      if (!g) { g = {head: null, rows: []}; byKey.set(k, g); groups.push(g); }
      g.rows.push(tr);
      if (tr.dataset.grphead === '1' && g.head === null) g.head = tr;
    });
    groups.forEach(g => {
      let v = g.head ? cellOf(g.head) : '';
      if (v === '') {
        const hit = g.rows.find(tr => cellOf(tr) !== '');
        v = hit ? cellOf(hit) : '';
      }
      g.key = v;
    });
    groups.sort((a, b) => cmp(a.key, b.key));
    groups.forEach(g => g.rows.forEach(tr => tbody.appendChild(tr)));
  } else {
    rows.sort((a, b) => cmp(cellOf(a), cellOf(b)));
    rows.forEach(r => tbody.appendChild(r));
  }
  regroupDates(sid);
}

// ── Cell tooltip: hover preview, click to pin ──
let _tipPinned = false;
let _tipEl = null;
function _ensureTip() {
  if (!_tipEl) {
    _tipEl = document.createElement('div');
    _tipEl.className = 'cell-tip';
    _tipEl.innerHTML = '<span class="tip-close" onclick="_hideTip()">✕</span><span class="tip-text"></span>';
    document.body.appendChild(_tipEl);
  }
  return _tipEl;
}
function _positionTip(tip, x, y) {
  tip.style.left = (x + 14) + 'px';
  tip.style.top = (y + 18) + 'px';
  requestAnimationFrame(() => {
    const r = tip.getBoundingClientRect();
    const vw = window.innerWidth, vh = window.innerHeight;
    if (r.right > vw - 8) tip.style.left = Math.max(4, vw - r.width - 8) + 'px';
    if (r.bottom > vh - 8) tip.style.top = Math.max(4, y - r.height - 14) + 'px';
  });
}
function _showTip(td, x, y) {
  const full = td.dataset.full;
  if (!full) return;
  const tip = _ensureTip();
  tip.querySelector('.tip-text').textContent = full;
  tip.style.display = 'block';
  _positionTip(tip, x, y);
}
function _hideTip() {
  if (_tipEl) _tipEl.style.display = 'none';
  if (_tipEl) _tipEl.classList.remove('pinned');
  _tipPinned = false;
}
document.addEventListener('mouseover', e => {
  if (_tipPinned) return;
  const td = e.target.closest('td[data-full], th[data-full]');
  if (td) _showTip(td, e.clientX, e.clientY);
});
document.addEventListener('mousemove', e => {
  if (_tipPinned || !_tipEl || _tipEl.style.display !== 'block') return;
  const td = e.target.closest('td[data-full], th[data-full]');
  if (td) _positionTip(_tipEl, e.clientX, e.clientY);
});
document.addEventListener('mouseout', e => {
  if (_tipPinned) return;
  const td = e.target.closest('td[data-full], th[data-full]');
  if (td) _hideTip();
});
document.addEventListener('click', e => {
  const td = e.target.closest('td[data-full], th[data-full]');
  if (td) {
    e.stopPropagation();
    if (_tipPinned) { _hideTip(); return; }
    _tipPinned = true;
    _showTip(td, e.clientX, e.clientY);
    _tipEl.classList.add('pinned');
  } else if (!e.target.closest('.cell-tip')) {
    _hideTip();
  }
});

// ── 行クリックで選択状態を固定する (客先要望 2026-08-05) ──
// セルのツールチップ側 (document の click) が stopPropagation するため,
// より内側の tbody に委譲して先に処理させる。
// もう一度同じ行をクリックすると解除。1つの表につき1行だけ選択する。
function toggleCompact(btn) {
  const panel = btn.closest('.panel');
  const on = panel.classList.toggle('compact');
  btn.textContent = on ? btn.dataset.off : btn.dataset.on;
}

function initRowSelect() {
  document.querySelectorAll('.data-tbl tbody').forEach(tbody => {
    tbody.addEventListener('click', e => {
      // 詳細へのリンク等の操作は邪魔しない
      if (e.target.closest('a, button, input, select')) return;
      const tr = e.target.closest('tr');
      if (!tr || tr.classList.contains('date-band')) return;
      const already = tr.classList.contains('row-selected');
      tbody.querySelectorAll('tr.row-selected')
           .forEach(x => x.classList.remove('row-selected'));
      if (!already) tr.classList.add('row-selected');
    });
  });
}

// init
showTab('01');
regroupDates('02');
initRowSelect();
"""


def write_html(sheets: dict, out_path: str, lang: str = "ja") -> None:
    """sheets: {'01': {header, rows, notices?}, '02': ..., ...}"""
    i18n = I18N.get(lang, I18N["ja"])
    tab_labels = i18n["tabs"]

    tabs_html = "".join(
        f'<button class="tab-btn" data-tab="{sid}" onclick="showTab(\'{sid}\')">{label}</button>'
        for sid, label in tab_labels.items()
        if sid in sheets
    )

    panels_html = "".join(
        _render_sheet(sid, tab_labels.get(sid, sid), sheets[sid], i18n)
        for sid in tab_labels
        if sid in sheets
    )

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    font  = i18n["font"]
    css   = CSS.replace(
        '"Hiragino Kaku Gothic ProN","Yu Gothic",Meiryo,"Noto Sans JP",sans-serif',
        font,
    )

    html = f"""<!doctype html>
<html lang="{i18n['lang_attr']}">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>{i18n['title']}</title>
<style>{css}</style>
</head>
<body>
<div class="app-header">
  <div class="app-title">{i18n['title']}</div>
  <div class="app-meta">{i18n['generated']}: {stamp}</div>
  <div class="tabs">{tabs_html}</div>
</div>
<div class="main">
  {panels_html}
</div>
<script>{JS}</script>
</body>
</html>"""

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)


def read_excel_and_write_html(xlsx_path: str, html_path: str, lang: str = "ja") -> None:
    """既存の Excel ファイルを読んで HTML を生成する."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets: dict = {}

    sheet_map = {
        "01_一次承認チェック": "01",
        "02_二次承認詳細": "02",
        "03_差異一覧": "03",
        "04_差戻し文面候補": "04",
        "05_取込ログ": "05",
        "06_判定ルール": "06",
        "07_マスタ確認": "07",
    }

    for sheet_name, sid in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        if sid == "01":
            notices = []
            header_idx = 0
            for i, row in enumerate(all_rows[:6]):
                v = row[0]
                if v and str(v).startswith("・"):
                    notices.append(str(v))
                elif any(cell == "伝票No." for cell in row):
                    header_idx = i
                    break
            header = [str(v) if v is not None else "" for v in all_rows[header_idx]]
            data_rows = [
                [str(v) if v is not None else "" for v in r]
                for r in all_rows[header_idx + 1:]
                if any(v is not None for v in r)
            ]
            sheets[sid] = {"header": header, "rows": data_rows, "notices": notices}
        else:
            header = [str(v) if v is not None else "" for v in all_rows[0]]
            data_rows = [
                [str(v) if v is not None else "" for v in r]
                for r in all_rows[1:]
                if any(v is not None for v in r)
            ]
            sheets[sid] = {"header": header, "rows": data_rows}

    write_html(sheets, html_path, lang=lang)
    print(f"HTML 出力 [{lang}]: {html_path}")


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python3 html_writer_web.py <excel_path> [html_path]")
        sys.exit(1)
    xlsx = sys.argv[1]
    html = sys.argv[2] if len(sys.argv) > 2 else xlsx.replace(".xlsx", ".html")
    read_excel_and_write_html(xlsx, html)
